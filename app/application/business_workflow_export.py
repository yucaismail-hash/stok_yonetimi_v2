"""Small, read-only Excel representation of one persisted Business Workflow."""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook


class BusinessWorkflowExportService:
    """Export only the durable aggregate envelope; never reruns analytics."""

    @staticmethod
    def build_workbook(aggregate: dict[str, Any]) -> BytesIO:
        coverage = aggregate.get("analysis_coverage") if isinstance(aggregate, dict) else None
        if not isinstance(coverage, dict):
            coverage = {"total_scope_count": 0, "fully_analyzed_count": 0, "partially_analyzed_count": 0, "excluded_count": 0, "exclusions": []}
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Özet"
        summary.append(["İşletme Analizi Kapsamı", "Değer"])
        summary.append(["Toplam ürün", coverage.get("total_scope_count", 0)])
        summary.append(["Tam analiz", coverage.get("fully_analyzed_count", 0)])
        summary.append(["Kısmi analiz", coverage.get("partially_analyzed_count", 0)])
        summary.append(["Analiz edilmeyen", coverage.get("excluded_count", 0)])
        summary.append(["Kapsam modu", coverage.get("scope_mode", "LATEST_UPLOAD")])
        summary.append(["Son yüklemede bulunan", coverage.get("latest_upload_count", 0)])
        summary.append(["Son yüklemede olmayan", coverage.get("absent_from_latest_upload_count", 0)])
        summary.append(["Güncel snapshot uyarısı", coverage.get("current_snapshot_warning_count", 0)])
        summary.append(["Önceki yüklemeden taşınan master uyarısı", coverage.get("stale_master_warning_count", 0)])
        exclusions = workbook.create_sheet("Analiz_Edilmeyenler")
        exclusions.append(["Ürün kodu", "Ürün adı", "Modül", "Durum", "Açıklama", "Mevcut hafta", "Gerekli hafta", "Son gözlem dönemi"])
        for item in coverage.get("exclusions", []):
            if not isinstance(item, dict):
                continue
            exclusions.append([
                item.get("material_code"), item.get("product_name"), item.get("capability"), item.get("status"),
                item.get("message"), item.get("available_weeks"), item.get("required_weeks"), item.get("latest_observation_period"),
            ])
        simulation = aggregate.get("simulation") if isinstance(aggregate, dict) else None
        simulation_items = simulation.get("items") if isinstance(simulation, dict) else None
        if isinstance(simulation_items, list) and simulation_items:
            incoming = workbook.create_sheet("Gelen_Arz")
            incoming.append(["Ürün kodu", "Kullanılan güncel stok", "Kullanılan gelen arz", "Planlanan teslim", "Gelen arz durumu", "Snapshot durumu", "Yeniden sipariş ufku (gün)", "Net gereksinim"])
            for item in simulation_items:
                if not isinstance(item, dict):
                    continue
                incoming.append([
                    item.get("material_code"), item.get("initial_stock"), item.get("incoming_supply_qty_used"),
                    item.get("incoming_supply_delivery_date"), item.get("incoming_supply_status"),
                    item.get("open_order_snapshot_state"), item.get("replenishment_horizon_days"), item.get("net_requirement_after_incoming_supply"),
                ])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
