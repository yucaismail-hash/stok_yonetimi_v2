"""FU2 staged canonical XLSX ingestion; Excel is never an analytical authority."""
from __future__ import annotations

import hashlib
import io
import json
import re
from datetime import date, datetime, timezone

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.application.actual_weekly_ledger import ActualWeeklyLedgerService
from app.models.actuals import ActualWeeklyRevision
from app.models.dataset import Dataset, DatasetEvent, DatasetState, DatasetValidationResult, DatasetVersion
from app.models.dataset_version_product_input import DatasetVersionProductInput
from app.services.dataset.ingestion_policy import EVENT_TYPES, validate_demand_type, validate_service_level
from app.services.dataset.weekly_normalization import parse_weekly_period
from app.services.security import EncryptionService


SHEET = "Talep_Gecmisi"
OFFICIAL_V3_SHEET = "Temel_Veriler"
OFFICIAL_V3_HEADERS = ("Ürün Kodu", "Ürün Adı", "Ürün Grubu", "Ürün Sınıfı", "Ürün Seviyesi", "Dönem Başı Stok", "Tedarik Süresi (Gün)", "Sipariş Parti Büyüklüğü", "Birim Maliyet (TL)", "Stok Tutma Oranı (%)", "Stok Tükenme Maliyeti")
REQUIRED = ("Malzeme Kodu", "Talep Tipi", "Ürün Seviyesi", "Dönem", "Miktar")
OPTIONAL = ("Ürün Grubu", "Ürün Sınıfı")
ALIASES = {
    "Malzeme Kodu": {"malzeme kodu", "ürün kodu", "product code", "material code"},
    "Talep Tipi": {"talep tipi", "demand type"},
    "Ürün Seviyesi": {"ürün seviyesi", "product level", "material type"},
    "Dönem": {"dönem", "period", "week"},
    "Miktar": {"miktar", "quantity", "demand"},
    "Ürün Grubu": {"ürün grubu", "product group"},
    "Ürün Sınıfı": {"ürün sınıfı", "product class"},
}
LEVELS = {"finished_good", "semi_finished_good", "raw_material"}
SUPPLIER_MAPPING_HEADERS = ("Ürün Kodu", "Tedarikçi Kodu", "Tedarik Payı (%)", "Açık Sipariş", "Planlanan Teslim Tarihi")
SUPPLIER_HEADERS = ("Tedarikçi Kodu", "Tedarikçi Adı", "Sipariş Karşılama Oranı (%)", "Terminden Önce Teslim (%)", "Termininde Teslim (%)", "Terminden Sonra Teslim (%)", "Ortalama Teslim Süresi (Gün)", "Teslim Süresi Std Sapma")
EVENT_HEADERS = ("Yıl", "Başlangıç Hafta", "Bitiş Hafta", "Ürün Grubu", "Ürün Sınıfı (Opsiyonel)", "Event Tipi", "Etki Değeri (%) (Opsiyonel)", "Referans Ürün Grubu (Opsiyonel)", "Referans Ürün Sınıfı (Opsiyonel)", "Açıklama (Opsiyonel)")
_ISO_HEADER = re.compile(r"^\d{4}-W\d{2}$")


class CanonicalExcelError(ValueError):
    pass


def _norm(value):
    return " ".join(str(value or "").strip().casefold().split())


def _issue(code, sheet, row, column, severity, message):
    return {"code": code, "sheet": sheet, "row": row, "column": column, "severity": severity, "message": message}


def _blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _number(value, *, sheet, row, column, errors, required, minimum=None, strictly_positive=False):
    if _blank(value):
        if required:
            errors.append(_issue("REQUIRED_VALUE_MISSING", sheet, row, column, "ERROR", f"{column} zorunludur."))
        return None
    if isinstance(value, bool):
        errors.append(_issue("INVALID_NUMERIC_VALUE", sheet, row, column, "ERROR", f"{column} sayısal olmalıdır."))
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        errors.append(_issue("INVALID_NUMERIC_VALUE", sheet, row, column, "ERROR", f"{column} sayısal olmalıdır."))
        return None
    if (strictly_positive and numeric <= 0) or (minimum is not None and numeric < minimum):
        message = f"{column} sıfırdan büyük olmalıdır." if strictly_positive else f"{column} {minimum} değerinden küçük olamaz."
        errors.append(_issue("VALUE_OUT_OF_RANGE", sheet, row, column, "ERROR", message))
        return None
    return numeric


def template_bytes() -> bytes:
    book = Workbook(); sheet = book.active; sheet.title = OFFICIAL_V3_SHEET
    weeks = [f"2026-W{week:02d}" for week in range(1, 53)]
    sheet.append(list(OFFICIAL_V3_HEADERS) + weeks)
    sheet.append(["SKU-001", "Örnek Ürün", "Örnek Grup", "Örnek Sınıf", "Mamul", 250, 7, 50, 125.5, 0.02, 500] + [40] * 52)
    for name, headers in (("Malzeme_Tedarikciler", ("Ürün Kodu", "Tedarikçi Kodu", "Tedarik Payı (%)", "Açık Sipariş", "Planlanan Teslim Tarihi")), ("Tedarikciler", ("Tedarikçi Kodu", "Tedarikçi Adı", "Sipariş Karşılama Oranı (%)", "Terminden Önce Teslim (%)", "Termininde Teslim (%)", "Terminden Sonra Teslim (%)", "Ortalama Teslim Süresi (Gün)", "Teslim Süresi Std Sapma")), ("Events", ("Yıl", "Başlangıç Hafta", "Bitiş Hafta", "Ürün Grubu", "Ürün Sınıfı (Opsiyonel)", "Event Tipi", "Etki Değeri (%) (Opsiyonel)", "Referans Ürün Grubu (Opsiyonel)", "Referans Ürün Sınıfı (Opsiyonel)", "Açıklama (Opsiyonel)")), ("Data_Requirement_Matrix", ("Alan", "Durum", "Kaynak", "Not"))):
        tab = book.create_sheet(name); tab.append(list(headers))
    output = io.BytesIO(); book.save(output); return output.getvalue()


def parse_workbook(content: bytes):
    """Backward-compatible public parser returning canonical rows and blocking issues."""
    rows, errors, _, _ = parse_workbook_details(content)
    return rows, errors


def parse_workbook_details(content: bytes):
    try:
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise CanonicalExcelError("WORKBOOK_UNREADABLE") from exc
    if OFFICIAL_V3_SHEET in book.sheetnames:
        return _parse_official_v3(book)
    if SHEET not in book.sheetnames:
        return [], [_issue("REQUIRED_SHEET_MISSING", SHEET, None, None, "ERROR", "Talep_Gecmisi sayfası bulunamadı.")], [], {}
    sheet = book[SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], [_issue("SHEET_EMPTY", SHEET, None, None, "ERROR", "Talep_Gecmisi sayfası boş.")], [], {}
    header = list(rows[0]); mapped = {}; errors = []
    for index, value in enumerate(header):
        matches = [canonical for canonical, aliases in ALIASES.items() if _norm(value) in aliases]
        if len(matches) > 1:
            errors.append({"code":"AMBIGUOUS_COLUMN","sheet":SHEET,"row":1,"column":str(value),"severity":"ERROR","message":"Kolon eşlemesi belirsiz."})
        elif matches:
            if matches[0] in mapped:
                errors.append({"code":"AMBIGUOUS_COLUMN","sheet":SHEET,"row":1,"column":str(value),"severity":"ERROR","message":"Aynı kanonik alan iki kolondan geliyor."})
            mapped[matches[0]] = index
    for field in REQUIRED:
        if field not in mapped:
            errors.append({"code":"REQUIRED_COLUMN_MISSING","sheet":SHEET,"row":1,"column":field,"severity":"ERROR","message":f"Zorunlu kolon eksik: {field}."})
    if errors: return [], errors, [], {}
    parsed = []; identities = set()
    for row_no, values in enumerate(rows[1:], 2):
        if not any(value not in (None, "") for value in values): continue
        def cell(field): return values[mapped[field]] if mapped[field] < len(values) else None
        item = {"material_code": str(cell("Malzeme Kodu") or "").strip(), "demand_type": str(cell("Talep Tipi") or "").strip().lower(), "product_level": str(cell("Ürün Seviyesi") or "").strip().lower(), "period": str(cell("Dönem") or "").strip(), "quantity": cell("Miktar"), "product_group": str(cell("Ürün Grubu") or "").strip() if "Ürün Grubu" in mapped else None, "product_class": str(cell("Ürün Sınıfı") or "").strip() if "Ürün Sınıfı" in mapped else None}
        try:
            if not item["material_code"]: raise CanonicalExcelError("MATERIAL_CODE_REQUIRED")
            item["demand_type"] = validate_demand_type(item["demand_type"])
            if item["demand_type"] not in {"sales", "consumption"}: raise CanonicalExcelError("DEMAND_TYPE_UNSUPPORTED")
            if item["product_level"] not in LEVELS: raise CanonicalExcelError("PRODUCT_LEVEL_UNSUPPORTED")
            item["period"] = parse_weekly_period(item["period"]).period
            if isinstance(item["quantity"], bool) or float(item["quantity"]) < 0: raise CanonicalExcelError("QUANTITY_INVALID")
            item["quantity"] = float(item["quantity"])
            identity = (item["material_code"], item["demand_type"], item["period"])
            if identity in identities: raise CanonicalExcelError("DUPLICATE_ROW_IDENTITY")
            identities.add(identity); parsed.append(item)
        except (ValueError, TypeError, CanonicalExcelError) as exc:
            raw_code = str(exc)
            code = {
                "invalid demand_type": "DEMAND_TYPE_UNSUPPORTED",
                "weekly period must use YYYY-Www": "PERIOD_INVALID",
                "invalid ISO weekly period": "PERIOD_INVALID",
            }.get(raw_code, raw_code)
            errors.append({"code":code,"sheet":SHEET,"row":row_no,"column":None,"severity":"ERROR","message":"Geçersiz satır; malzeme, talep tipi, ürün seviyesi, dönem ve miktarı kontrol edin."})
    if not parsed and not errors:
        errors.append({"code":"SHEET_EMPTY","sheet":SHEET,"row":None,"column":None,"severity":"ERROR","message":"Talep_Gecmisi sayfasında veri yok."})
    return parsed, errors, [], {}


def _parse_official_v3(book):
    sheet = book[OFFICIAL_V3_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], [_issue("SHEET_EMPTY", OFFICIAL_V3_SHEET, None, None, "ERROR", "Temel_Veriler sayfası boş.")], [], {}
    header = list(rows[0])
    indexes = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
    errors, warnings = [], []
    for field in OFFICIAL_V3_HEADERS:
        if field not in indexes:
            errors.append(_issue("REQUIRED_COLUMN_MISSING", OFFICIAL_V3_SHEET, 1, field, "ERROR", f"Zorunlu kolon eksik: {field}."))
    periods = []
    for value in header:
        text = str(value).strip() if value is not None else ""
        if _ISO_HEADER.fullmatch(text):
            try:
                periods.append(parse_weekly_period(text))
            except ValueError:
                errors.append(_issue("INVALID_ISO_WEEK", OFFICIAL_V3_SHEET, 1, text, "ERROR", "Hafta başlığı geçerli ISO YYYY-Www olmalıdır."))
    periods.sort(key=lambda item: (item.year, item.week))
    if not periods:
        errors.append(_issue("WEEKLY_COLUMNS_MISSING", OFFICIAL_V3_SHEET, 1, None, "ERROR", "En az bir ISO haftalık talep kolonu gereklidir."))
    if errors:
        return [], errors, warnings, {}

    level_map = {"mamul":"finished_good", "yarı mamul":"semi_finished_good", "yari mamul":"semi_finished_good", "hammadde":"raw_material", **{value:value for value in LEVELS}}
    parsed, identities = [], set()
    missing_economics = {"Birim Maliyet (TL)": 0, "Stok Tutma Oranı (%)": 0, "Stok Tükenme Maliyeti": 0}
    for row_no, values in enumerate(rows[1:], 2):
        if not any(not _blank(value) for value in values):
            continue
        row_errors = []
        def cell(name):
            return values[indexes[name]] if indexes[name] < len(values) else None
        code = str(cell("Ürün Kodu") or "").strip()
        group = str(cell("Ürün Grubu") or "").strip()
        raw_level = cell("Ürün Seviyesi")
        if not code:
            row_errors.append(_issue("REQUIRED_VALUE_MISSING", OFFICIAL_V3_SHEET, row_no, "Ürün Kodu", "ERROR", "Ürün Kodu zorunludur."))
        if not group:
            row_errors.append(_issue("REQUIRED_VALUE_MISSING", OFFICIAL_V3_SHEET, row_no, "Ürün Grubu", "ERROR", "Ürün Grubu zorunludur."))
        level = level_map.get(_norm(raw_level))
        if _blank(raw_level):
            row_errors.append(_issue("REQUIRED_VALUE_MISSING", OFFICIAL_V3_SHEET, row_no, "Ürün Seviyesi", "ERROR", "Ürün Seviyesi zorunludur."))
        elif level is None:
            row_errors.append(_issue("INVALID_ENUM_VALUE", OFFICIAL_V3_SHEET, row_no, "Ürün Seviyesi", "ERROR", "Geçersiz Ürün Seviyesi. İzin verilen değerler: Mamul, Yarı Mamul, Hammadde."))
        initial_stock = _number(cell("Dönem Başı Stok"), sheet=OFFICIAL_V3_SHEET, row=row_no, column="Dönem Başı Stok", errors=row_errors, required=True, minimum=0)
        lead_time = _number(cell("Tedarik Süresi (Gün)"), sheet=OFFICIAL_V3_SHEET, row=row_no, column="Tedarik Süresi (Gün)", errors=row_errors, required=True, strictly_positive=True)
        lot_size = _number(cell("Sipariş Parti Büyüklüğü"), sheet=OFFICIAL_V3_SHEET, row=row_no, column="Sipariş Parti Büyüklüğü", errors=row_errors, required=True, minimum=0)
        economics = {}
        for field, target in (("Birim Maliyet (TL)", "unit_cost"), ("Stok Tutma Oranı (%)", "holding_rate"), ("Stok Tükenme Maliyeti", "stockout_cost")):
            value = cell(field)
            if _blank(value):
                missing_economics[field] += 1
                economics[target] = None
            else:
                economics[target] = _number(value, sheet=OFFICIAL_V3_SHEET, row=row_no, column=field, errors=row_errors, required=False, minimum=0)
        metadata = {"material_code": code, "product_name": str(cell("Ürün Adı") or "").strip() or None, "product_group": group or None, "product_class": str(cell("Ürün Sınıfı") or "").strip() or None, "product_level": level, "initial_stock": initial_stock, "lead_time_days": lead_time, "lot_size": lot_size, **economics}
        row_points = []
        for period in periods:
            value = cell(period.period)
            if _blank(value):
                continue
            quantity = _number(value, sheet=OFFICIAL_V3_SHEET, row=row_no, column=period.period, errors=row_errors, required=False, minimum=0)
            if quantity is None:
                continue
            identity = (code, period.period)
            if identity in identities:
                row_errors.append(_issue("DUPLICATE_ROW_IDENTITY", OFFICIAL_V3_SHEET, row_no, period.period, "ERROR", "Aynı ürün ve hafta için birden fazla talep değeri bulunuyor."))
            else:
                identities.add(identity)
                row_points.append({**metadata, "period": period.period, "quantity": quantity})
        if not row_errors:
            parsed.extend(row_points)
        errors.extend(row_errors)
    for field, count in missing_economics.items():
        if count:
            capability = "Maliyet/kazanım çıktıları sınırlı olacaktır." if field == "Birim Maliyet (TL)" else "İlgili ekonomik çıktılar sınırlı olacaktır."
            warnings.append(_issue("OPTIONAL_VALUE_MISSING", OFFICIAL_V3_SHEET, None, field, "WARNING", f"{field}, {count} üründe bulunmuyor. {capability}"))
    supplier_inputs, supplier_errors = _parse_optional_supplier_sheets(book, {row["material_code"] for row in parsed})
    event_inputs, event_errors = _parse_optional_events(book, {row["product_group"] for row in parsed if row["product_group"]}, {(row["product_group"], row["product_class"]) for row in parsed if row["product_group"] and row["product_class"]})
    errors.extend(supplier_errors); errors.extend(event_errors)
    return parsed, errors, warnings, {"v3_supplier_inputs": supplier_inputs, "v3_events": event_inputs}


def _optional_sheet_rows(book, name, headers, aliases=()):
    if name not in book.sheetnames:
        return [], {}, []
    rows = list(book[name].iter_rows(values_only=True))
    if not rows:
        return [], {}, []
    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    indexes = {value: index for index, value in enumerate(header) if value}
    for source, target in aliases:
        if source in indexes and target not in indexes:
            indexes[target] = indexes[source]
    data = [(row_no, values) for row_no, values in enumerate(rows[1:], 2) if any(not _blank(value) for value in values)]
    return data, indexes, header


def _parse_optional_supplier_sheets(book, material_codes):
    errors = []
    supplier_data, supplier_indexes, supplier_header = _optional_sheet_rows(book, "Tedarikciler", SUPPLIER_HEADERS, (("Teslim Süresi Std. Sapması", "Teslim Süresi Std Sapma"),))
    mapping_data, mapping_indexes, mapping_header = _optional_sheet_rows(book, "Malzeme_Tedarikciler", SUPPLIER_MAPPING_HEADERS)
    suppliers = {}
    if supplier_data:
        for field in ("Tedarikçi Kodu",):
            if field not in supplier_indexes:
                errors.append(_issue("REQUIRED_COLUMN_MISSING", "Tedarikciler", 1, field, "ERROR", f"Zorunlu kolon eksik: {field}."))
        for row_no, values in supplier_data:
            def cell(name): return values[supplier_indexes[name]] if name in supplier_indexes and supplier_indexes[name] < len(values) else None
            code = str(cell("Tedarikçi Kodu") or "").strip()
            if not code:
                errors.append(_issue("REQUIRED_VALUE_MISSING", "Tedarikciler", row_no, "Tedarikçi Kodu", "ERROR", "Tedarikçi Kodu zorunludur.")); continue
            metrics = {}
            for field in ("Sipariş Karşılama Oranı (%)", "Terminden Önce Teslim (%)", "Termininde Teslim (%)", "Terminden Sonra Teslim (%)"):
                value = cell(field)
                if not _blank(value): metrics[field] = _number(value, sheet="Tedarikciler", row=row_no, column=field, errors=errors, required=False, minimum=0)
                if field in metrics and metrics[field] is not None and metrics[field] > 1:
                    errors.append(_issue("VALUE_OUT_OF_RANGE", "Tedarikciler", row_no, field, "ERROR", f"{field} 0 ile 1 arasında olmalıdır."))
            for field in ("Ortalama Teslim Süresi (Gün)", "Teslim Süresi Std Sapma"):
                value = cell(field)
                if not _blank(value): metrics[field] = _number(value, sheet="Tedarikciler", row=row_no, column=field, errors=errors, required=False, minimum=0)
            suppliers[code] = {"supplier_code": code, "supplier_name": str(cell("Tedarikçi Adı") or "").strip() or None, "metrics": metrics}
    mappings = []
    if mapping_data:
        for field in ("Ürün Kodu", "Tedarikçi Kodu", "Tedarik Payı (%)"):
            if field not in mapping_indexes:
                errors.append(_issue("REQUIRED_COLUMN_MISSING", "Malzeme_Tedarikciler", 1, field, "ERROR", f"Zorunlu kolon eksik: {field}."))
        for row_no, values in mapping_data:
            def cell(name): return values[mapping_indexes[name]] if name in mapping_indexes and mapping_indexes[name] < len(values) else None
            material = str(cell("Ürün Kodu") or "").strip(); supplier = str(cell("Tedarikçi Kodu") or "").strip()
            if not material: errors.append(_issue("REQUIRED_VALUE_MISSING", "Malzeme_Tedarikciler", row_no, "Ürün Kodu", "ERROR", "Ürün Kodu zorunludur."))
            elif material not in material_codes: errors.append(_issue("INVALID_REFERENCE", "Malzeme_Tedarikciler", row_no, "Ürün Kodu", "ERROR", f"`{material}` kodlu malzeme `Temel_Veriler` tablosunda bulunamadı. Bu malzemeyi `Temel_Veriler` tablosuna ekleyin veya ilgili satırı `Malzeme_Tedarikciler` tablosundan silin."))
            if not supplier: errors.append(_issue("REQUIRED_VALUE_MISSING", "Malzeme_Tedarikciler", row_no, "Tedarikçi Kodu", "ERROR", "Tedarikçi Kodu zorunludur."))
            elif supplier not in suppliers: errors.append(_issue("INVALID_REFERENCE", "Malzeme_Tedarikciler", row_no, "Tedarikçi Kodu", "ERROR", f"`{supplier}` kodlu tedarikçi `Tedarikciler` tablosunda bulunamadı. Bu tedarikçiyi `Tedarikciler` tablosuna ekleyin veya ilgili satırı `Malzeme_Tedarikciler` tablosundan silin."))
            share = _number(cell("Tedarik Payı (%)"), sheet="Malzeme_Tedarikciler", row=row_no, column="Tedarik Payı (%)", errors=errors, required=True, minimum=0)
            if share is not None and share > 1: errors.append(_issue("VALUE_OUT_OF_RANGE", "Malzeme_Tedarikciler", row_no, "Tedarik Payı (%)", "ERROR", "Tedarik Payı (%) 0 ile 1 arasında olmalıdır."))
            open_order = _number(cell("Açık Sipariş"), sheet="Malzeme_Tedarikciler", row=row_no, column="Açık Sipariş", errors=errors, required=False, minimum=0)
            planned = cell("Planlanan Teslim Tarihi")
            if not _blank(planned):
                try:
                    if isinstance(planned, datetime): planned = planned.date()
                    if isinstance(planned, str): planned = date.fromisoformat(planned)
                    if not isinstance(planned, date): raise ValueError
                except ValueError:
                    errors.append(_issue("INVALID_DATE_VALUE", "Malzeme_Tedarikciler", row_no, "Planlanan Teslim Tarihi", "ERROR", "Planlanan Teslim Tarihi geçerli bir tarih olmalıdır."))
            mappings.append({"material_code": material, "supplier_code": supplier, "share": share, "open_order": open_order, "planned_delivery_date": planned.isoformat() if isinstance(planned, date) else None})
    return {"suppliers": list(suppliers.values()), "material_suppliers": mappings}, errors


def _parse_optional_events(book, groups, group_classes):
    data, indexes, header = _optional_sheet_rows(book, "Events", EVENT_HEADERS)
    if not data:
        return [], []
    errors, events = [], []
    for field in ("Yıl", "Başlangıç Hafta", "Bitiş Hafta", "Ürün Grubu", "Event Tipi"):
        if field not in indexes: errors.append(_issue("REQUIRED_COLUMN_MISSING", "Events", 1, field, "ERROR", f"Zorunlu kolon eksik: {field}."))
    for row_no, values in data:
        def cell(name): return values[indexes[name]] if name in indexes and indexes[name] < len(values) else None
        year = _number(cell("Yıl"), sheet="Events", row=row_no, column="Yıl", errors=errors, required=True, minimum=1)
        start = _number(cell("Başlangıç Hafta"), sheet="Events", row=row_no, column="Başlangıç Hafta", errors=errors, required=True, minimum=1)
        end = _number(cell("Bitiş Hafta"), sheet="Events", row=row_no, column="Bitiş Hafta", errors=errors, required=True, minimum=1)
        if all(value is not None for value in (year, start, end)):
            try:
                parse_weekly_period(f"{int(year):04d}-W{int(start):02d}")
            except ValueError: errors.append(_issue("INVALID_ISO_WEEK", "Events", row_no, "Başlangıç Hafta", "ERROR", "Başlangıç Hafta geçerli ISO hafta değeri olmalıdır."))
            try:
                parse_weekly_period(f"{int(year):04d}-W{int(end):02d}")
            except ValueError: errors.append(_issue("INVALID_ISO_WEEK", "Events", row_no, "Bitiş Hafta", "ERROR", "Bitiş Hafta geçerli ISO hafta değeri olmalıdır."))
            if start > end: errors.append(_issue("VALUE_OUT_OF_RANGE", "Events", row_no, "Bitiş Hafta", "ERROR", "Bitiş Hafta, Başlangıç Hafta değerinden önce olamaz."))
        group = str(cell("Ürün Grubu") or "").strip()
        if not group: errors.append(_issue("REQUIRED_VALUE_MISSING", "Events", row_no, "Ürün Grubu", "ERROR", "Ürün Grubu zorunludur."))
        elif group not in groups: errors.append(_issue("INVALID_REFERENCE", "Events", row_no, "Ürün Grubu", "ERROR", f"`{group}` Ürün Grubu `Temel_Veriler` tablosunda bulunamadı. Önce ilgili Ürün Grubu'nu ekleyin veya Events satırını düzeltin."))
        product_class = str(cell("Ürün Sınıfı (Opsiyonel)") or "").strip() or None
        if product_class and (group, product_class) not in group_classes: errors.append(_issue("INVALID_REFERENCE", "Events", row_no, "Ürün Sınıfı (Opsiyonel)", "ERROR", f"`{product_class}` Ürün Sınıfı, `{group}` Ürün Grubu altında `Temel_Veriler` tablosunda bulunamadı. Ürün Sınıfı'nı düzeltin veya Temel_Veriler'e ekleyin."))
        event_type = str(cell("Event Tipi") or "").strip()
        if event_type not in EVENT_TYPES: errors.append(_issue("INVALID_ENUM_VALUE", "Events", row_no, "Event Tipi", "ERROR", "Geçersiz Event Tipi."))
        impact = cell("Etki Değeri (%) (Opsiyonel)")
        if not _blank(impact): _number(impact, sheet="Events", row=row_no, column="Etki Değeri (%) (Opsiyonel)", errors=errors, required=False)
        events.append({"year": int(year) if year is not None else None, "start_week": int(start) if start is not None else None, "end_week": int(end) if end is not None else None, "product_group": group or None, "product_class": product_class, "event_type": event_type or None})
    return events, errors


class CanonicalExcelIngestionService:
    def get_current_accepted(self, session: Session, company_id):
        """Return the tenant's most recently accepted active dataset deterministically."""
        row = (
            session.query(Dataset, DatasetEvent)
            .join(
                DatasetEvent,
                (DatasetEvent.dataset_id == Dataset.id) & (DatasetEvent.event_type == "accepted"),
            )
            .filter(
                Dataset.company_id == company_id,
                Dataset.is_active.is_(True),
                Dataset.is_deleted.is_(False),
                Dataset.state == DatasetState.APPROVED,
                DatasetEvent.is_deleted.is_(False),
            )
            .order_by(DatasetEvent.created_at.desc(), DatasetEvent.id.desc())
            .first()
        )
        if row is None:
            return None
        dataset, accepted_event = row
        return {
            "dataset_id": str(dataset.id),
            "status": "READY_FOR_WORKFLOW",
            "accepted": True,
            "accepted_at": accepted_event.created_at,
            "created_at": dataset.created_at,
            "source_name": dataset.source_name,
            "record_count": dataset.record_count,
            "material_count": dataset.sku_count,
        }

    def stage(self, session: Session, company_id, user_id, filename: str, content: bytes, demand_type=None, service_level=None):
        if not filename.lower().endswith(".xlsx"):
            raise CanonicalExcelError("FILE_TYPE_INVALID")
        rows, errors, warnings, extras = parse_workbook_details(content)
        is_v3 = OFFICIAL_V3_SHEET in load_workbook(io.BytesIO(content), read_only=True).sheetnames
        if is_v3:
            try: demand_type = validate_demand_type(demand_type)
            except ValueError: demand_type = None
            if demand_type not in {"sales", "consumption"}: errors.append({"code":"DEMAND_TYPE_REQUIRED","sheet":OFFICIAL_V3_SHEET,"row":None,"column":None,"severity":"ERROR","message":"Talep tipi Wizard/API metadata olarak zorunludur."})
            try: service_level = validate_service_level(service_level or {"mode":"automatic"})
            except ValueError: errors.append({"code":"SERVICE_LEVEL_INVALID","sheet":OFFICIAL_V3_SHEET,"row":None,"column":None,"severity":"ERROR","message":"Servis seviyesi otomatik veya 0-1 arası manuel olmalıdır."}); service_level = {"mode":"automatic"}
            if demand_type: rows=[{**row,"demand_type":demand_type} for row in rows]
        semantic = json.dumps({"contract":"official_v3","demand_type":demand_type,"service_level":service_level}, sort_keys=True, separators=(",", ":")).encode() if is_v3 else b""
        fingerprint = hashlib.sha256(str(company_id).encode() + semantic + content).hexdigest()
        existing = session.query(Dataset).filter_by(company_id=company_id, dataset_hash=fingerprint, is_active=True).one_or_none()
        if existing:
            return existing, True
        periods = [item["period"] for item in rows]
        contract = "official_v3" if is_v3 else "fu2_weekly_v1"
        payload = {"actual_rows": rows, "contract":contract, "demand_type":demand_type, "service_level":service_level or {"mode":"automatic"}, **extras}
        dataset = Dataset(company_id=company_id, user_id=user_id, uploaded_by=user_id, dataset_hash=fingerprint, source_type="excel", source_name=filename, state=DatasetState.VALIDATED if not errors else DatasetState.FAILED, record_count=len(rows), sku_count=len({item["material_code"] for item in rows}), encrypted_data=EncryptionService(session).encrypt_dataset(user_id, payload), is_active=True)
        session.add(dataset); session.flush()
        session.add(DatasetValidationResult(dataset_id=dataset.id, is_valid=not errors, errors=errors, warnings=warnings, validated_by=user_id, requires_user_approval=not errors))
        session.add(DatasetEvent(dataset_id=dataset.id, event_type="validated" if not errors else "validation_failed", event_data={"periods": sorted(periods), "contract":contract, "demand_type":demand_type, "service_level":service_level}, created_by=user_id))
        session.commit(); return dataset, False

    def accept(self, session: Session, company_id, user_id, dataset_id):
        dataset = session.query(Dataset).filter_by(id=dataset_id, company_id=company_id, user_id=user_id, is_active=True).one_or_none()
        if not dataset: raise CanonicalExcelError("DATASET_UNAVAILABLE")
        if dataset.state == DatasetState.APPROVED: return {"status":"READY_FOR_WORKFLOW","dataset_id":str(dataset.id),"idempotent":True}
        validation = session.query(DatasetValidationResult).filter_by(dataset_id=dataset.id).order_by(DatasetValidationResult.validated_at.desc()).first()
        if not validation or not validation.is_valid: raise CanonicalExcelError("DATASET_NOT_READY_FOR_ACCEPTANCE")
        payload = EncryptionService(session).decrypt_dataset(user_id, dataset.encrypted_data)
        try:
            version = DatasetVersion(dataset_id=dataset.id, version_number=1, dataset_hash=dataset.dataset_hash, record_count=dataset.record_count, sku_count=dataset.sku_count, created_by=user_id, is_current=True)
            session.add(version); session.flush()
            if payload.get("contract") == "official_v3":
                inputs = {row["material_code"]: row for row in payload["actual_rows"]}
                session.add_all([DatasetVersionProductInput(company_id=company_id, dataset_version_id=version.id, material_code=row["material_code"], product_name=row.get("product_name"), product_group=row.get("product_group"), product_class=row.get("product_class"), product_level=row["product_level"], initial_stock=row["initial_stock"], lead_time_days=row["lead_time_days"], lot_size=row["lot_size"], unit_cost=row["unit_cost"], holding_rate=row["holding_rate"], stockout_cost=row["stockout_cost"]) for row in inputs.values()])
                session.flush()
            grouped = {}
            for row in payload["actual_rows"]: grouped.setdefault(row["demand_type"], []).append(row)
            ledger = ActualWeeklyLedgerService()
            summary = {kind: ledger.ingest_dataset_actuals_in_session(session, company_id, user_id, dataset.id, rows, kind) for kind, rows in grouped.items()}
            proposed = session.query(ActualWeeklyRevision).filter_by(company_id=company_id, source_dataset_id=dataset.id, approval_status="proposed").all()
            for revision in proposed: ledger.approve_revision_in_session(session, company_id, revision.id, user_id)
            dataset.state = DatasetState.APPROVED
            session.add(DatasetEvent(dataset_id=dataset.id, event_type="accepted", event_data={"ledger":summary,"status":"READY_FOR_WORKFLOW","contract":payload.get("contract")}, created_by=user_id))
            session.commit()
            return {"status":"READY_FOR_WORKFLOW","dataset_id":str(dataset.id),"version_id":str(version.id),"ledger":summary,"idempotent":False}
        except Exception:
            session.rollback()
            raise
