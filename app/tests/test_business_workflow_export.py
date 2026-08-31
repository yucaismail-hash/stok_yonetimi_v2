import unittest

from openpyxl import load_workbook

from app.application.business_workflow_export import BusinessWorkflowExportService
from app.application.business_workflow_presentation import BusinessWorkflowPresentationService


class BusinessWorkflowExportTests(unittest.TestCase):
    coverage = {
        "total_scope_count": 3, "fully_analyzed_count": 2,
        "partially_analyzed_count": 1, "excluded_count": 0,
        "exclusions": [{
            "material_code": "SKU-C", "product_name": "Kısa Geçmiş", "capability": "backtest",
            "status": "EXCLUDED", "message": "Bu ürün için en az 16 haftalık geçmiş veri gerekir. Mevcut geçmiş: 7 hafta.",
            "available_weeks": 7, "required_weeks": 16, "latest_observation_period": "2026-W07",
        }],
        "scope_mode": "ALL_ACTIVE_SKUS", "latest_upload_count": 1,
        "absent_from_latest_upload_count": 2, "current_snapshot_warning_count": 2,
        "stale_master_warning_count": 1,
    }

    def test_coverage_summary_and_exclusions_sheet_are_exported(self):
        workbook = BusinessWorkflowExportService.build_workbook({
            "analysis_coverage": self.coverage,
        })
        book = load_workbook(workbook)
        self.assertEqual(book.sheetnames, ["Özet", "Analiz_Edilmeyenler"])
        self.assertEqual(book["Özet"]["B2"].value, 3)
        self.assertEqual(book["Özet"]["B6"].value, "ALL_ACTIVE_SKUS")
        self.assertEqual(book["Özet"]["B10"].value, 1)
        self.assertEqual(book["Analiz_Edilmeyenler"]["A2"].value, "SKU-C")
        self.assertEqual(book["Analiz_Edilmeyenler"]["G2"].value, 16)

    def test_presentation_exposes_only_valid_persisted_coverage(self):
        aggregate = type("Aggregate", (), {"inline_result": {"analysis_coverage": self.coverage}})()
        coverage = BusinessWorkflowPresentationService._coverage_view(aggregate)
        self.assertEqual(coverage.total_scope_count, 3)
        self.assertEqual(coverage.exclusions[0]["material_code"], "SKU-C")

    def test_current_incoming_supply_evidence_is_exported_from_persisted_simulation(self):
        workbook = BusinessWorkflowExportService.build_workbook({
            "analysis_coverage": self.coverage,
            "simulation": {"items": [{
                "material_code": "SKU-A", "initial_stock": 100, "incoming_supply_qty_used": 50,
                "incoming_supply_delivery_date": "2026-01-05", "incoming_supply_status": "VALID_WITHIN_REPLENISHMENT_HORIZON",
                "open_order_snapshot_state": "SNAPSHOT_SUPPLIED", "net_requirement_after_incoming_supply": 25,
            }]},
        })
        book = load_workbook(workbook)
        self.assertEqual(book["Gelen_Arz"]["C2"].value, 50)
        self.assertEqual(book["Gelen_Arz"]["H2"].value, 25)


if __name__ == "__main__":
    unittest.main()
