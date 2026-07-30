# app/utils/excel_exporter.py
# STOKONOMI EXECUTIVE EXCEL REPORT V2
# Profesyonel Yönetim Raporu - 8 Sayfa

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Any, Optional
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import (
    PieChart, PieChart3D, BarChart, BarChart3D,
    Reference, Series, LineChart
)
from openpyxl.chart.label import DataLabelList


class ExcelExporter:
    """STOKONOMI Executive Excel Report V2"""
    
    def __init__(self):
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.report_date = datetime.now().strftime('%d.%m.%Y %H:%M')
        
        # Kurumsal Renkler - ARGB Formatında
        self.COLORS = {
            'primary': 'FF1f4e79',
            'secondary': 'FF2e75b6',
            'success': 'FF2e7d32',
            'warning': 'FFed6c02',
            'danger': 'FFd32f2f',
            'info': 'FF1976d2',
            'light_gray': 'FFF5F5F5',
            'medium_gray': 'FFE0E0E0',
            'white': 'FFFFFFFF',
            'gold': 'FFFFD700',
        }
        
        self.AI_DECISION_MAP = {
            'increase_safety_stock': 'Emniyet Stoğunu Artır',
            'decrease_safety_stock': 'Emniyet Stoğunu Azalt',
            'maintain_current': 'Mevcut Politikayı Koru',
            'urgent_action': 'Acil Müdahale Gerekli',
            'review_supplier': 'Tedarikçiyi İncele',
            'change_forecast_model': 'Tahmin Modelini Değiştir',
            'investigate_variability': 'Talep Değişkenliğini Araştır',
            'seasonal_adjustment': 'Mevsimsel Ayarlama Yap',
            'normal_monitoring': 'Normal Takip'
        }
        
        self.METHOD_MAP = {
            'classic_ss': 'Klasik SS',
            'croston_ss': 'Croston SS',
            'syntetos_boylan_ss': 'SB Croston SS',
            'bootstrapping_ss': 'Bootstrapping SS',
            'ml_ss': 'ML Tabanlı SS',
            'hybrid_ss': 'Hibrit SS (AI Önerilen)'
        }
    
    # ============================================================
    # 📌 ANA METOD
    # ============================================================
    
    def export_bulk_report(
        self, 
        materials_data: List[Dict], 
        learning_rules: List[Dict] = None,
        ai_decision: Dict = None,
        executive_summary: Dict = None
    ) -> io.BytesIO:
        """Executive Excel Raporu - 8 Sayfa"""
        
        for m in materials_data:
            if 'description' not in m:
                m['description'] = m.get('product_name', m.get('material_name', ''))
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            self._create_executive_dashboard(writer, materials_data, executive_summary, ai_decision)
            self._create_ai_decisions_sheet(writer, materials_data)
            self._create_all_calculations_sheet(writer, materials_data)
            self._create_technical_analysis_sheet(writer, materials_data)
            self._create_ai_explanations_sheet(writer, materials_data)
            self._create_learning_memory_sheet(writer, learning_rules)
            self._create_critical_products_sheet(writer, materials_data)
            self._create_raw_data_sheet(writer, materials_data)
        
        output.seek(0)
        return output
    
    # ============================================================
    # 📄 SAYFA 1: Yönetici Özeti
    # ============================================================
    
    def _create_executive_dashboard(
        self, 
        writer: pd.ExcelWriter, 
        materials_data: List[Dict],
        executive_summary: Dict = None,
        ai_decision: Dict = None
    ):
        """Sayfa 1: Executive Dashboard"""
        
        if not materials_data:
            df_empty = pd.DataFrame({'Bilgi': ['Henüz analiz verisi yok']})
            df_empty.to_excel(writer, sheet_name='Yönetici Özeti', index=False)
            return
        
        total = len(materials_data)
        high_risk = len([m for m in materials_data if m.get('risk_score', 0) > 0.5])
        medium_risk = len([m for m in materials_data if 0.3 < m.get('risk_score', 0) <= 0.5])
        low_risk = len([m for m in materials_data if m.get('risk_score', 0) <= 0.3])
        
        increase_count = 0
        decrease_count = 0
        maintain_count = 0
        urgent_count = 0
        
        for m in materials_data:
            decision = m.get('ai_decision', {}).get('decision', '')
            if decision == 'increase_safety_stock':
                increase_count += 1
            elif decision == 'decrease_safety_stock':
                decrease_count += 1
            elif decision == 'maintain_current':
                maintain_count += 1
            elif decision == 'urgent_action':
                urgent_count += 1
        
        method_counts = {}
        for m in materials_data:
            method = m.get('recommended_method', 'hybrid_ss')
            method_counts[method] = method_counts.get(method, 0) + 1
        
        abc_counts = {'A': 0, 'B': 0, 'C': 0}
        xyz_counts = {'X': 0, 'Y': 0, 'Z': 0}
        
        for m in materials_data:
            abc = m.get('abc', 'C')
            xyz = m.get('xyz', 'Z')
            abc_counts[abc] = abc_counts.get(abc, 0) + 1
            xyz_counts[xyz] = xyz_counts.get(xyz, 0) + 1
        
        total_ss = sum(m.get('recommended_value', 0) for m in materials_data)
        avg_risk = np.mean([m.get('risk_score', 0) for m in materials_data]) if materials_data else 0
        
        # KPI Kartları
        kpi_data = [
            {'KPI': 'Analiz Edilen Ürün', 'Değer': total, 'Birim': 'Adet'},
            {'KPI': 'Önerilen Toplam SS', 'Değer': f"{total_ss:,.0f}", 'Birim': 'Birim'},
            {'KPI': 'Ortalama Risk', 'Değer': f"{avg_risk:.2f}", 'Birim': ''},
            {'KPI': 'Kritik Ürün', 'Değer': high_risk, 'Birim': 'Adet'},
            {'KPI': 'AI Artır Önerisi', 'Değer': increase_count, 'Birim': 'Adet'},
            {'KPI': 'AI Koru Önerisi', 'Değer': maintain_count, 'Birim': 'Adet'},
            {'KPI': 'AI Azalt Önerisi', 'Değer': decrease_count, 'Birim': 'Adet'},
        ]
        
        df_kpi = pd.DataFrame(kpi_data)
        df_kpi.to_excel(writer, sheet_name='Yönetici Özeti', index=False, startrow=0)
        
        workbook = writer.book
        worksheet = writer.sheets['Yönetici Özeti']
        
        # 1. AI Karar Dağılımı
        if increase_count + decrease_count + maintain_count + urgent_count > 0:
            chart_data = [
                ['Karar', 'Adet'],
                ['Artır', increase_count],
                ['Koru', maintain_count],
                ['Azalt', decrease_count],
                ['Acil', urgent_count],
            ]
            for row_idx, row_data in enumerate(chart_data, start=1):
                for col_idx, val in enumerate(row_data, start=7):
                    worksheet.cell(row=row_idx, column=col_idx, value=val)
            
            pie = PieChart()
            pie.title = "AI Karar Dağılımı"
            pie.width = 12
            pie.height = 8
            data = Reference(worksheet, min_col=8, min_row=2, max_row=5)
            labels = Reference(worksheet, min_col=7, min_row=2, max_row=5)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            worksheet.add_chart(pie, "M2")
        
        # 2. Metot Dağılımı
        if method_counts:
            method_data = [['Metot', 'Adet']]
            method_labels = {
                'classic_ss': 'Klasik',
                'croston_ss': 'Croston',
                'syntetos_boylan_ss': 'SB Croston',
                'bootstrapping_ss': 'Bootstrap',
                'ml_ss': 'ML',
                'hybrid_ss': 'Hibrit'
            }
            for method, count in method_counts.items():
                method_data.append([method_labels.get(method, method), count])
            
            for row_idx, row_data in enumerate(method_data, start=1):
                for col_idx, val in enumerate(row_data, start=7):
                    worksheet.cell(row=row_idx + 6, column=col_idx, value=val)
            
            bar = BarChart()
            bar.title = "Önerilen Metot Dağılımı"
            bar.width = 12
            bar.height = 8
            bar.type = "bar"
            bar.gapWidth = 50
            data = Reference(worksheet, min_col=8, min_row=7, max_row=7+len(method_counts))
            labels = Reference(worksheet, min_col=7, min_row=8, max_row=7+len(method_counts))
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(labels)
            worksheet.add_chart(bar, "M20")
        
        # 3. Risk Dağılımı
        risk_data = [
            ['Risk', 'Adet'],
            ['Düşük', low_risk],
            ['Orta', medium_risk],
            ['Yüksek', high_risk],
        ]
        for row_idx, row_data in enumerate(risk_data, start=1):
            for col_idx, val in enumerate(row_data, start=12):
                worksheet.cell(row=row_idx, column=col_idx, value=val)
        
        risk_bar = BarChart()
        risk_bar.title = "Risk Dağılımı"
        risk_bar.width = 10
        risk_bar.height = 6
        data = Reference(worksheet, min_col=13, min_row=2, max_row=4)
        labels = Reference(worksheet, min_col=12, min_row=2, max_row=4)
        risk_bar.add_data(data, titles_from_data=True)
        risk_bar.set_categories(labels)
        worksheet.add_chart(risk_bar, "M38")
        
        # 4. ABC Dağılımı
        abc_data = [
            ['ABC', 'Adet'],
            ['A', abc_counts.get('A', 0)],
            ['B', abc_counts.get('B', 0)],
            ['C', abc_counts.get('C', 0)],
        ]
        for row_idx, row_data in enumerate(abc_data, start=1):
            for col_idx, val in enumerate(row_data, start=12):
                worksheet.cell(row=row_idx + 6, column=col_idx, value=val)
        
        abc_bar = BarChart()
        abc_bar.title = "ABC Dağılımı"
        abc_bar.width = 10
        abc_bar.height = 6
        data = Reference(worksheet, min_col=13, min_row=7, max_row=9)
        labels = Reference(worksheet, min_col=12, min_row=7, max_row=9)
        abc_bar.add_data(data, titles_from_data=True)
        abc_bar.set_categories(labels)
        worksheet.add_chart(abc_bar, "V38")
        
        # 5. XYZ Dağılımı
        xyz_data = [
            ['XYZ', 'Adet'],
            ['X', xyz_counts.get('X', 0)],
            ['Y', xyz_counts.get('Y', 0)],
            ['Z', xyz_counts.get('Z', 0)],
        ]
        for row_idx, row_data in enumerate(xyz_data, start=1):
            for col_idx, val in enumerate(row_data, start=12):
                worksheet.cell(row=row_idx + 12, column=col_idx, value=val)
        
        xyz_bar = BarChart()
        xyz_bar.title = "XYZ Dağılımı"
        xyz_bar.width = 10
        xyz_bar.height = 6
        data = Reference(worksheet, min_col=13, min_row=13, max_row=15)
        labels = Reference(worksheet, min_col=12, min_row=13, max_row=15)
        xyz_bar.add_data(data, titles_from_data=True)
        xyz_bar.set_categories(labels)
        worksheet.add_chart(xyz_bar, "V50")
        
        # 6. Toplam SS Karşılaştırması
        classic_total = sum(m.get('classic_ss', 0) for m in materials_data)
        croston_total = sum(m.get('croston_ss', 0) for m in materials_data)
        sb_total = sum(m.get('syntetos_boylan_ss', 0) for m in materials_data)
        bootstrap_total = sum(m.get('bootstrapping_ss', 0) for m in materials_data)
        ml_total = sum(m.get('ml_ss', 0) for m in materials_data)
        hybrid_total = sum(m.get('hybrid_ss', 0) for m in materials_data)
        
        ss_comparison = [
            ['Metot', 'Toplam SS'],
            ['Klasik', classic_total],
            ['Croston', croston_total],
            ['SB Croston', sb_total],
            ['Bootstrap', bootstrap_total],
            ['ML', ml_total],
            ['Hibrit (AI)', hybrid_total],
        ]
        
        for row_idx, row_data in enumerate(ss_comparison, start=1):
            for col_idx, val in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx + 25, column=col_idx + 12, value=val)
        
        ss_bar = BarChart()
        ss_bar.title = "Toplam Emniyet Stoğu Karşılaştırması"
        ss_bar.width = 12
        ss_bar.height = 8
        data = Reference(worksheet, min_col=14, min_row=26, max_row=31)
        labels = Reference(worksheet, min_col=13, min_row=26, max_row=31)
        ss_bar.add_data(data, titles_from_data=True)
        ss_bar.set_categories(labels)
        worksheet.add_chart(ss_bar, "M55")
        
        # Stiller
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, color='FFFFFFFF', size=11)
                cell.fill = PatternFill(start_color=self.COLORS['primary'], end_color=self.COLORS['primary'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 18
        worksheet.column_dimensions['C'].width = 12
    
    # ============================================================
    # 📄 SAYFA 2: AI Kararları
    # ============================================================
    
    def _create_ai_decisions_sheet(self, writer: pd.ExcelWriter, materials_data: List[Dict]):
        """Sayfa 2: AI Kararları - Kullanıcı Dili"""
        
        if not materials_data:
            pd.DataFrame({'Bilgi': ['Henüz AI kararı yok']}).to_excel(
                writer, sheet_name='AI Kararları', index=False
            )
            return
        
        rows = []
        for m in materials_data:
            ai = m.get('ai_decision', {})
            decision_raw = ai.get('decision', 'normal_monitoring')
            confidence = self._calculate_dynamic_confidence(m)
            expected_impact = self._calculate_expected_impact(m, ai)
            review_days = self._calculate_review_days(m)
            
            rows.append({
                'Malzeme Kodu': m.get('material_code', ''),
                'Ürün Adı': m.get('description', ''),
                'AI Kararı': self.AI_DECISION_MAP.get(decision_raw, decision_raw),
                'Karar Nedeni': ' | '.join(ai.get('reasons', ['Analiz sonucu'])),
                'Risk Açıklaması': self._get_risk_description(m),
                'Beklenen Etki': expected_impact,
                'Önerilen Aksiyon': self._get_action_text(m, ai),
                'Güven Skoru': f"%{int(confidence * 100)}",
                'Sonraki İnceleme': f"{review_days} gün"
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='AI Kararları', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['AI Kararları']
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = PatternFill(start_color=self.COLORS['primary'], end_color=self.COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 35
    
    # ============================================================
    # 📄 SAYFA 3: Tüm Hesaplamalar
    # ============================================================
    
    def _create_all_calculations_sheet(self, writer: pd.ExcelWriter, materials_data: List[Dict]):
        """Sayfa 3: Tüm Hesaplamalar"""
        
        if not materials_data:
            pd.DataFrame({'Bilgi': ['Henüz veri yok']}).to_excel(
                writer, sheet_name='Tüm Hesaplamalar', index=False
            )
            return
        
        rows = []
        for m in materials_data:
            confidence = self._calculate_dynamic_confidence(m)
            rows.append({
                'Malzeme Kodu': m.get('material_code', ''),
                'Ürün Adı': m.get('description', ''),
                'Grup': m.get('group', ''),
                'ABC': m.get('abc', '-'),
                'XYZ': m.get('xyz', '-'),
                'CV': round(m.get('cv', 0), 4),
                'Pattern': m.get('pattern_label', ''),
                'Lead Time': m.get('lead_time_days', '-'),
                'Classic SS': round(m.get('classic_ss', 0), 0),
                'Croston SS': round(m.get('croston_ss', 0), 0),
                'SB Croston SS': round(m.get('syntetos_boylan_ss', 0), 0),
                'Bootstrap SS': round(m.get('bootstrapping_ss', 0), 0),
                'ML SS': round(m.get('ml_ss', 0), 0),
                'Hybrid SS': round(m.get('hybrid_ss', 0), 0),
                'AI Seçtiği Metot': self.METHOD_MAP.get(m.get('recommended_method', 'hybrid_ss'), m.get('recommended_method', '-')),
                'Önerilen SS': round(m.get('recommended_value', 0), 0),
                'Risk Skoru': round(m.get('risk_score', 0), 3),
                'Confidence': f"%{int(confidence * 100)}"
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Tüm Hesaplamalar', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Tüm Hesaplamalar']
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = PatternFill(start_color=self.COLORS['primary'], end_color=self.COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 35
    
    # ============================================================
    # 📄 SAYFA 4: Teknik Analiz
    # ============================================================
    
    def _create_technical_analysis_sheet(self, writer: pd.ExcelWriter, materials_data: List[Dict]):
        """Sayfa 4: Teknik Analiz"""
        
        if not materials_data:
            pd.DataFrame({'Bilgi': ['Henüz veri yok']}).to_excel(
                writer, sheet_name='Teknik Analiz', index=False
            )
            return
        
        rows = []
        for m in materials_data:
            rows.append({
                'Malzeme Kodu': m.get('material_code', ''),
                'Ürün Adı': m.get('description', ''),
                'CV': round(m.get('cv', 0), 4),
                'Pattern': m.get('pattern_label', ''),
                'Trend': m.get('trend_direction', ''),
                'Forecast Model': m.get('forecast_model_label', ''),
                'Seasonality': m.get('seasonality_label', ''),
                'Zero Ratio': round(m.get('zero_ratio', 0), 4),
                'Lead Time': m.get('lead_time_days', '-'),
                'ABC': m.get('abc', '-'),
                'XYZ': m.get('xyz', '-'),
                'Risk': m.get('risk_level', ''),
                'Intermittent': m.get('intermittent_level', '')
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Teknik Analiz', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Teknik Analiz']
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = PatternFill(start_color=self.COLORS['secondary'], end_color=self.COLORS['secondary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 35
    
    # ============================================================
    # 📄 SAYFA 5: AI Açıklamaları
    # ============================================================
    
    def _create_ai_explanations_sheet(self, writer: pd.ExcelWriter, materials_data: List[Dict]):
        """Sayfa 5: AI Açıklamaları"""
        
        if not materials_data:
            pd.DataFrame({'Bilgi': ['Henüz AI açıklaması yok']}).to_excel(
                writer, sheet_name='AI Açıklamaları', index=False
            )
            return
        
        rows = []
        for m in materials_data:
            ai = m.get('ai_decision', {})
            rows.append({
                'Malzeme Kodu': m.get('material_code', ''),
                'Ürün Adı': m.get('description', ''),
                'AI Açıklaması': self._generate_detailed_explanation(m, ai),
                'Önerilen Aksiyon': self._get_action_text(m, ai),
                'Risk Seviyesi': m.get('risk_level', 'Düşük')
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='AI Açıklamaları', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['AI Açıklamaları']
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = PatternFill(start_color=self.COLORS['primary'], end_color=self.COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 35
        worksheet.column_dimensions['C'].width = 70
    
    # ============================================================
    # 📄 SAYFA 6: Learning Engine
    # ============================================================
    
    def _create_learning_memory_sheet(self, writer: pd.ExcelWriter, learning_rules: List[Dict] = None):
        """Sayfa 6: Learning Engine"""
        
        if not learning_rules:
            df_empty = pd.DataFrame({
                'Bilgi': ['📚 Learning Engine Henüz Aktif Değil'],
                'Açıklama': ['Analiz yaptıkça AI işletmenizi tanımaya başlayacak.']
            })
            df_empty.to_excel(writer, sheet_name='Learning Engine', index=False)
            return
        
        rows = []
        for rule in learning_rules:
            rows.append({
                'Kural ID': rule.get('rule_id', ''),
                'Kural Adı': rule.get('rule_name', ''),
                'Tip': rule.get('rule_type', '').upper(),
                'Açıklama': rule.get('description', ''),
                'Güven Skoru': f"%{int(rule.get('confidence_score', 0) * 100)}",
                'Kullanım': rule.get('usage_count', 0),
                'Doğrulandı': '✅' if rule.get('is_verified') else '⏳'
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Learning Engine', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Learning Engine']
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = PatternFill(start_color=self.COLORS['secondary'], end_color=self.COLORS['secondary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # ============================================================
    # 📄 SAYFA 7: Kritik Ürünler
    # ============================================================
    
    def _create_critical_products_sheet(self, writer: pd.ExcelWriter, materials_data: List[Dict]):
        """Sayfa 7: Kritik Ürünler"""
        
        critical_items = [m for m in materials_data if m.get('risk_score', 0) > 0.5]
        
        if not critical_items:
            pd.DataFrame({'Bilgi': ['🎉 Kritik ürün bulunmuyor.']}).to_excel(
                writer, sheet_name='Kritik Ürünler', index=False
            )
            return
        
        rows = []
        for m in critical_items:
            rows.append({
                'Malzeme Kodu': m.get('material_code', ''),
                'Ürün Adı': m.get('description', ''),
                'Grup': m.get('group', ''),
                'Risk Skoru': round(m.get('risk_score', 0), 3),
                'Risk Seviyesi': m.get('risk_level', ''),
                'CV': round(m.get('cv', 0), 4),
                'Pattern': m.get('pattern_label', ''),
                'Önerilen SS': round(m.get('recommended_value', 0), 0),
                'Mevcut SS': round(m.get('classic_ss', 0), 0),
                'AI Kararı': self.AI_DECISION_MAP.get(
                    m.get('ai_decision', {}).get('decision', ''),
                    m.get('ai_decision', {}).get('decision', 'İnceleniyor')
                )
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Kritik Ürünler', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Kritik Ürünler']
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = PatternFill(start_color=self.COLORS['danger'], end_color=self.COLORS['danger'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 35
    
    # ============================================================
    # 📄 SAYFA 8: Ham Veri
    # ============================================================
    
    def _create_raw_data_sheet(self, writer: pd.ExcelWriter, materials_data: List[Dict]):
        """Sayfa 8: Ham Veri"""
        
        if not materials_data:
            pd.DataFrame({'Bilgi': ['Ham veri yok']}).to_excel(
                writer, sheet_name='Ham Veri', index=False
            )
            return
        
        raw_data = []
        for m in materials_data:
            row = dict(m)
            if 'ai_decision' in row and isinstance(row['ai_decision'], dict):
                row['ai_decision'] = str(row['ai_decision'])
            raw_data.append(row)
        
        df = pd.DataFrame(raw_data)
        df.to_excel(writer, sheet_name='Ham Veri', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Ham Veri']
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = PatternFill(start_color=self.COLORS['medium_gray'], end_color=self.COLORS['medium_gray'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # ============================================================
    # 📌 YARDIMCI FONKSİYONLAR
    # ============================================================
    
    def _calculate_dynamic_confidence(self, material: Dict) -> float:
        confidence = 0.5
        cv = material.get('cv', 0)
        if cv < 0.3:
            confidence += 0.3
        elif cv < 0.6:
            confidence += 0.15
        else:
            confidence += 0.05
        
        pattern = material.get('pattern', 'DEGISKEN')
        if pattern in ['DUZENLI_SABIT', 'DUZENLI_ARTS', 'DUZENLI_AZALIS']:
            confidence += 0.2
        elif pattern in ['DEGISKEN', 'YUKSEK_DEGISKEN']:
            confidence += 0.1
        else:
            confidence += 0.05
        
        data_len = len(material.get('historical_demand', []))
        if data_len >= 52:
            confidence += 0.2
        elif data_len >= 26:
            confidence += 0.15
        elif data_len >= 13:
            confidence += 0.1
        else:
            confidence += 0.05
        
        zero_ratio = material.get('zero_ratio', 0)
        if zero_ratio < 0.1:
            confidence += 0.1
        elif zero_ratio < 0.3:
            confidence += 0.05
        
        lt = material.get('lead_time_days', 14)
        if lt <= 14:
            confidence += 0.1
        elif lt <= 21:
            confidence += 0.05
        
        risk = material.get('risk_score', 0)
        if risk <= 0.3:
            confidence += 0.1
        elif risk <= 0.5:
            confidence += 0.05
        
        return min(0.98, max(0.2, confidence))
    
    def _calculate_expected_impact(self, material: Dict, ai_decision: Dict) -> str:
        risk = material.get('risk_score', 0)
        decision = ai_decision.get('decision', 'maintain_current')
        if decision == 'increase_safety_stock':
            return "Stok tükenme riski %40 azalır, stok maliyeti %15 artar" if risk > 0.5 else "Stok tükenme riski %25 azalır, stok maliyeti %8 artar"
        elif decision == 'decrease_safety_stock':
            return "Stok maliyeti %12 azalır, servis seviyesi %2 düşer"
        elif decision == 'urgent_action':
            return "Acil müdahale ile stok riski %50 azaltılabilir"
        else:
            return "Mevcut politika ile risk yönetilebilir seviyede"
    
    def _calculate_review_days(self, material: Dict) -> int:
        abc = material.get('abc', 'C')
        risk = material.get('risk_score', 0)
        base_days = 7 if abc == 'A' else 14 if abc == 'B' else 30
        if risk > 0.5:
            base_days = max(3, base_days // 2)
        elif risk > 0.3:
            base_days = max(5, int(base_days * 0.7))
        return base_days
    
    def _get_risk_description(self, material: Dict) -> str:
        risk = material.get('risk_score', 0)
        cv = material.get('cv', 0)
        lt = material.get('lead_time_days', 14)
        desc = []
        if risk > 0.5:
            desc.append("🔴 Yüksek risk")
        elif risk > 0.3:
            desc.append("🟡 Orta risk")
        else:
            desc.append("🟢 Düşük risk")
        if cv > 0.7:
            desc.append("Talep değişkenliği yüksek")
        elif cv > 0.4:
            desc.append("Talep değişkenliği orta")
        if lt > 21:
            desc.append("Lead Time uzun")
        return " | ".join(desc) if desc else "Risk seviyesi normal"
    
    def _get_action_text(self, material: Dict, ai_decision: Dict) -> str:
        decision = ai_decision.get('decision', 'normal_monitoring')
        action_map = {
            'increase_safety_stock': 'SS seviyesini artır',
            'decrease_safety_stock': 'SS seviyesini azalt',
            'maintain_current': 'Mevcut politikayı sürdür',
            'urgent_action': 'Acil aksiyon al',
            'review_supplier': 'Tedarikçiyi gözden geçir',
            'change_forecast_model': 'Tahmin modelini değiştir',
            'investigate_variability': 'Talep değişkenliğini araştır',
            'seasonal_adjustment': 'Mevsimsel ayarlama yap',
            'normal_monitoring': 'Normal takip'
        }
        return action_map.get(decision, 'Analiz sonuçlarını değerlendir')
    
    def _generate_detailed_explanation(self, material: Dict, ai_decision: Dict) -> str:
        code = material.get('material_code', 'Bu ürün')
        cv = material.get('cv', 0)
        lt = material.get('lead_time_days', 14)
        abc = material.get('abc', 'C')
        xyz = material.get('xyz', 'Z')
        pattern = material.get('pattern_label', 'Değişken')
        risk = material.get('risk_score', 0)
        decision = ai_decision.get('decision', 'maintain_current')
        
        parts = []
        if cv > 0.7:
            parts.append(f"{code} yüksek değişkenlik göstermektedir (CV={cv:.2f})")
        elif cv > 0.4:
            parts.append(f"{code} orta düzeyde değişkenlik göstermektedir (CV={cv:.2f})")
        else:
            parts.append(f"{code} düşük değişkenlik göstermektedir (CV={cv:.2f})")
        
        if lt > 21:
            parts.append(f"Lead Time uzun ({lt} gün) olduğu için stok riski artmaktadır")
        elif lt > 14:
            parts.append(f"Lead Time orta seviyede ({lt} gün)")
        else:
            parts.append(f"Lead Time kabul edilebilir seviyede ({lt} gün)")
        
        if abc == 'A' and xyz == 'Z':
            parts.append(f"ABC={abc} ve XYZ={xyz} sınıfında olduğu için kritik öneme sahiptir")
        elif abc == 'A':
            parts.append(f"ABC={abc} sınıfında olduğu için yüksek önceliklidir")
        
        if pattern in ['Düzenli Sabit', 'Düzenli Artan']:
            parts.append(f"Talep deseni {pattern} olduğu için tahmin edilebilir")
        else:
            parts.append(f"Talep deseni {pattern} olduğu için dikkatli yönetim gerekir")
        
        decision_text = self.AI_DECISION_MAP.get(decision, decision)
        parts.append(f"AI {decision_text} önermektedir")
        
        method = material.get('recommended_method_label', 'Hibrit')
        if method:
            parts.append(f"Önerilen metot: {method}")
        
        return ". ".join(parts) + "."
    
    # ============================================================
    # 📌 ESKİ FONKSİYONLAR (Korunuyor)
    # ============================================================
    
    def export_recommendations(
        self, 
        material_code: str, 
        material_data: Dict,
        simulation_result: Dict, 
        ai_analysis: Dict,
        optimized_params: Dict
    ) -> io.BytesIO:
        """Eski fonksiyon - korunuyor"""
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_data = {
                'Rapor Tarihi': datetime.now().strftime('%d.%m.%Y %H:%M'),
                'Malzeme Kodu': material_code,
                'Malzeme Açıklaması': material_data.get('description', ''),
                'Malzeme Grubu': material_data.get('group', ''),
            }
            summary_df = pd.DataFrame([summary_data])
            summary_df.to_excel(writer, sheet_name='Özet', index=False)
        output.seek(0)
        return output
    
    def _create_summary_sheet(self, material_code: str, material_data: Dict,
                              simulation_result: Dict, ai_analysis: Dict,
                              optimized_params: Dict) -> Dict:
        """Özet sayfası oluştur (Eski)"""
        service_level = simulation_result.get('service_level_actual', 0)
        stockout_prob = np.mean(simulation_result.get('stockout_probability', [0]))
        cvar_95 = simulation_result.get('cvar_95', 0)
        tail_risk = simulation_result.get('tail_risk', 0)
        
        return {
            'Rapor Tarihi': datetime.now().strftime('%d.%m.%Y %H:%M'),
            'Malzeme Kodu': material_code,
            'Malzeme Açıklaması': material_data.get('description', ''),
            'Malzeme Grubu': material_data.get('group', ''),
            'Mevcut Başlangıç Stoku': material_data.get('initial_stock', 0),
            'Önerilen Başlangıç Stoku': optimized_params.get('recommended_initial_stock', 0),
            'Mevcut EOQ': material_data.get('eoq', 0),
            'Önerilen EOQ': optimized_params.get('optimal_eoq', 0),
            'AI Önerilen Safety Stock': ai_analysis.get('ai_ss', 0),
            'Önerilen Safety Stock': optimized_params.get('safety_stock', 0),
            'Önerilen ROP': optimized_params.get('optimal_rop', 0),
            'Servis Seviyesi (Gerçekleşen)': f"{service_level*100:.1f}%",
            'Servis Seviyesi (Hedef)': "95%",
            'Stok Tükenme Olasılığı': f"{stockout_prob*100:.1f}%",
            'CVaR95': round(cvar_95, 2),
            'Tail Risk': round(tail_risk, 2),
            'Patern': ai_analysis.get('pattern', ''),
            'CV (Değişim Katsayısı)': round(ai_analysis.get('cv', 0), 4),
            'Risk Seviyesi': optimized_params.get('risk_level', ''),
            'AI Çarpan': optimized_params.get('pattern_multiplier', 1.0)
        }
    
    def _create_detail_sheet(self, material_data: Dict, simulation_result: Dict,
                            ai_analysis: Dict, optimized_params: Dict) -> pd.DataFrame:
        """Detaylı analiz sayfası oluştur (Eski)"""
        data = []
        
        historical = material_data.get('historical_demand', [])
        if historical:
            non_zero = [d for d in historical if d > 0]
            data.append({
                'Kategori': 'Talep İstatistikleri',
                'Parametre': 'Ortalama Talep',
                'Değer': f"{np.mean(non_zero):.2f}" if non_zero else "0",
                'Birim': 'Adet/Hafta'
            })
            data.append({
                'Kategori': 'Talep İstatistikleri',
                'Parametre': 'Standart Sapma',
                'Değer': f"{np.std(non_zero):.2f}" if non_zero else "0",
                'Birim': 'Adet'
            })
            data.append({
                'Kategori': 'Talep İstatistikleri',
                'Parametre': 'CV (Değişim Katsayısı)',
                'Değer': f"{ai_analysis.get('cv', 0):.4f}",
                'Birim': ''
            })
            data.append({
                'Kategori': 'Talep İstatistikleri',
                'Parametre': 'Sıfır Talep Oranı',
                'Değer': f"{ai_analysis.get('zero_ratio', 0):.4f}",
                'Birim': ''
            })
        
        data.append({
            'Kategori': 'Tedarik Bilgileri',
            'Parametre': 'Lead Time',
            'Değer': f"{material_data.get('lead_time_days', 0)}",
            'Birim': 'Gün'
        })
        data.append({
            'Kategori': 'Tedarik Bilgileri',
            'Parametre': 'Lead Time Talep',
            'Değer': f"{optimized_params.get('lead_time_demand', 0):.2f}",
            'Birim': 'Adet'
        })
        
        if 'simulation_results' in simulation_result:
            sim = simulation_result['simulation_results']
            methods = {
                'Klasik SS': sim.get('classic_ss', 0),
                'Croston SS': sim.get('croston_ss', 0),
                'Syntetos-Boylan SS': sim.get('syntetos_boylan_ss', 0),
                'Bootstrapping SS': sim.get('bootstrapping_ss', 0),
                'ML SS': sim.get('ml_ss', 0),
                'Hybrid SS': sim.get('hybrid_ss', 0)
            }
            for method, value in methods.items():
                data.append({
                    'Kategori': 'SS Metodları',
                    'Parametre': method,
                    'Değer': f"{value:.2f}",
                    'Birim': 'Adet'
                })
        
        return pd.DataFrame(data)
    
    def _create_demand_sheet(self, material_data: Dict) -> pd.DataFrame:
        """Haftalık talep sayfası oluştur (Eski)"""
        historical = material_data.get('historical_demand', [])
        df = pd.DataFrame({
            'Hafta': list(range(1, len(historical) + 1)),
            'Talep': historical
        })
        return df
    
    def _create_simulation_sheet(self, simulation_result: Dict) -> pd.DataFrame:
        """Simülasyon sonuçları sayfası oluştur (Eski)"""
        sim = simulation_result.get('simulation_results', {})
        
        avg_stock = sim.get('avg_stock', [])
        stockout_prob = sim.get('stockout_probability', [])
        expected_shortage = sim.get('expected_shortage', [])
        
        if not avg_stock:
            return pd.DataFrame()
        
        df = pd.DataFrame({
            'Hafta': list(range(1, len(avg_stock) + 1)),
            'Ortalama Stok': avg_stock,
            'Stok Tükenme Olasılığı': stockout_prob,
            'Beklenen Açık': expected_shortage
        })
        return df
    
    def _create_supplier_sheet(self, material_data: Dict) -> pd.DataFrame:
        """Tedarikçi bilgileri sayfası oluştur (Eski)"""
        suppliers = material_data.get('suppliers', [])
        if not suppliers:
            return pd.DataFrame()
        
        rows = []
        for s in suppliers:
            rows.append({
                'Tedarikçi Kodu': s.get('supplier_id', ''),
                'Pay': f"{s.get('share', 0)*100:.1f}%",
                'Açık Bakiye': s.get('open_qty', 0)
            })
        
        return pd.DataFrame(rows)
    
    def _create_action_plan_sheet(self, material_data: Dict, optimized_params: Dict) -> pd.DataFrame:
        """Aksiyon planı sayfası oluştur (Eski)"""
        initial_stock = material_data.get('initial_stock', 0)
        recommended_stock = optimized_params.get('recommended_initial_stock', 0)
        current_eoq = material_data.get('eoq', 0)
        optimal_eoq = optimized_params.get('optimal_eoq', 0)
        safety_stock = optimized_params.get('safety_stock', 0)
        
        actions = []
        
        if recommended_stock > initial_stock:
            actions.append({
                'Sıra': 1,
                'Aksiyon': 'Başlangıç Stok Artırımı',
                'Mevcut': initial_stock,
                'Önerilen': recommended_stock,
                'Artış': recommended_stock - initial_stock,
                'Birim': 'Adet'
            })
        
        if optimal_eoq > current_eoq:
            actions.append({
                'Sıra': 2,
                'Aksiyon': 'EOQ Yükseltme',
                'Mevcut': current_eoq,
                'Önerilen': optimal_eoq,
                'Artış': optimal_eoq - current_eoq,
                'Birim': 'Adet'
            })
        
        if safety_stock > 0:
            actions.append({
                'Sıra': 3,
                'Aksiyon': 'Safety Stock Güncelleme',
                'Mevcut': material_data.get('initial_stock', 0) - optimized_params.get('lead_time_demand', 0),
                'Önerilen': safety_stock,
                'Artış': safety_stock - (material_data.get('initial_stock', 0) - optimized_params.get('lead_time_demand', 0)),
                'Birim': 'Adet'
            })
        
        if not actions:
            actions.append({
                'Sıra': 1,
                'Aksiyon': 'Mevcut Politika Yeterli',
                'Mevcut': 0,
                'Önerilen': 0,
                'Artış': 0,
                'Birim': ''
            })
        
        return pd.DataFrame(actions)