"""FU2-R2A: one invalid XLSX must stage but never become Actual authority."""
import io, sys
from pathlib import Path
from time import perf_counter
from uuid import uuid4
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from openpyxl import load_workbook
from app.application.canonical_excel_ingestion import CanonicalExcelError, CanonicalExcelIngestionService, template_bytes
from app.auth import RegisterRequest, register
from app.database import SessionLocal
from app.models import ActualWeeklyObservation, DatasetVersion, RuntimeExecution, ForecastVintage, DecisionSnapshot
from scripts.verify_fu2_excel_ingestion import cleanup

def main():
 cleanup(); s=SessionLocal(); key=uuid4().hex; start=perf_counter()
 try:
  owner=register(RegisterRequest(email=f'fu2r2a-{key}@example.test',password='pilot-password-1',full_name='Pilot',company_name=f'FU2-R2A-{key}'),db=s)
 finally:s.close()
 try:
  started=perf_counter(); book=load_workbook(io.BytesIO(template_bytes())); ws=book['Talep_Gecmisi']; ws['B2']='INVALID_DEMAND'; raw=io.BytesIO();book.save(raw); generation=perf_counter()-started
  started=perf_counter(); s=SessionLocal(); dataset,retry=CanonicalExcelIngestionService().stage(s,owner['company_id'],owner['user_id'],'invalid.xlsx',raw.getvalue()); dataset_id=dataset.id; issue=dataset.validations[-1].errors[0]; staging=perf_counter()-started;print(f'company_id={owner["company_id"]} dataset_id={dataset_id} status={dataset.state.value} issues=1 READY_FOR_ACCEPTANCE=False',flush=True);print(issue,flush=True);assert issue['code']=='DEMAND_TYPE_UNSUPPORTED' and issue['severity']=='ERROR';s.close()
  s=SessionLocal();assert s.query(DatasetVersion).filter_by(dataset_id=dataset_id).count()==0;assert s.query(ActualWeeklyObservation).filter_by(source_dataset_id=dataset_id).count()==0;s.close();print('pre_accept versions=0 actuals=0',flush=True)
  started=perf_counter();s=SessionLocal()
  try:
   CanonicalExcelIngestionService().accept(s,owner['company_id'],owner['user_id'],dataset_id);raise AssertionError('invalid dataset accepted')
  except CanonicalExcelError as exc:
   assert str(exc)=='DATASET_NOT_READY_FOR_ACCEPTANCE';print(f'accept_rejected={exc}',flush=True)
  finally:s.close()
  rejected=perf_counter()-started;s=SessionLocal();assert s.query(DatasetVersion).filter_by(dataset_id=dataset_id).count()==0;assert s.query(ActualWeeklyObservation).filter_by(source_dataset_id=dataset_id).count()==0;assert s.query(RuntimeExecution).filter_by(company_id=owner['company_id']).count()==0;assert s.query(ForecastVintage).filter_by(company_id=owner['company_id']).count()==0;assert s.query(DecisionSnapshot).filter_by(company_id=owner['company_id']).count()==0;s.close();print(f'fresh_post_rejection PASS generation={generation:.3f}s staging={staging:.3f}s reject={rejected:.3f}s',flush=True)
 finally:
  cleanup();print('FU2-R2A cleanup/residue PASS',flush=True)
if __name__=='__main__':main()
