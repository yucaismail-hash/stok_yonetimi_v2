"""FU2-R2B same-workbook retry safety probe."""
import io, sys
from pathlib import Path
from time import perf_counter
from uuid import uuid4
sys.path.insert(0,str(Path(__file__).resolve().parents[1]))
from openpyxl import load_workbook
from app.application.canonical_excel_ingestion import CanonicalExcelIngestionService, template_bytes
from app.auth import RegisterRequest, register
from app.database import SessionLocal
from app.models import ActualWeeklyObservation,ActualWeeklyRevision,Dataset,DatasetVersion,RuntimeExecution,ForecastVintage,DecisionSnapshot
from scripts.verify_fu2_excel_ingestion import cleanup
def main():
 cleanup(); key=uuid4().hex; s=SessionLocal()
 try: owner=register(RegisterRequest(email=f'fu2r2b-{key}@example.test',password='pilot-password-1',full_name='Pilot',company_name=f'FU2-R2B-{key}'),db=s)
 finally:s.close()
 try:
  wb=load_workbook(io.BytesIO(template_bytes()));ws=wb['Talep_Gecmisi'];ws.delete_rows(2,ws.max_row)
  for week in range(1,5):ws.append(['SKU-FU2-R2B','sales','finished_good',f'2026-W{week:02d}',100+week,'G','C'])
  out=io.BytesIO();wb.save(out);content=out.getvalue(); service=CanonicalExcelIngestionService();print(f'workbook_sha256={__import__("hashlib").sha256(content).hexdigest()}',flush=True)
  t=perf_counter();s=SessionLocal();d,retry=service.stage(s,owner['company_id'],owner['user_id'],'retry.xlsx',content);did=d.id;assert not retry and d.validations[-1].is_valid;print(f'first_stage dataset={did} fingerprint={d.dataset_hash} seconds={perf_counter()-t:.3f}',flush=True);s.close()
  t=perf_counter();s=SessionLocal();result=service.accept(s,owner['company_id'],owner['user_id'],did);assert result['status']=='READY_FOR_WORKFLOW';print(f'first_accept version={result["version_id"]} seconds={perf_counter()-t:.3f}',flush=True);s.close()
  s=SessionLocal();base=(s.query(Dataset).filter_by(company_id=owner['company_id']).count(),s.query(DatasetVersion).filter_by(dataset_id=did).count(),s.query(ActualWeeklyObservation).filter_by(company_id=owner['company_id']).count(),s.query(ActualWeeklyRevision).filter_by(company_id=owner['company_id']).count());s.close();print(f'fresh_baseline={base}',flush=True)
  t=perf_counter();s=SessionLocal();again,retry=service.stage(s,owner['company_id'],owner['user_id'],'retry.xlsx',content);assert retry and again.id==did;print(f'retry_stage dataset={again.id} status={again.state.value} seconds={perf_counter()-t:.3f}',flush=True);t=perf_counter();result2=service.accept(s,owner['company_id'],owner['user_id'],again.id);assert result2['idempotent'];print(f'retry_accept={result2["status"]} idempotent=True seconds={perf_counter()-t:.3f}',flush=True);s.close()
  s=SessionLocal();after=(s.query(Dataset).filter_by(company_id=owner['company_id']).count(),s.query(DatasetVersion).filter_by(dataset_id=did).count(),s.query(ActualWeeklyObservation).filter_by(company_id=owner['company_id']).count(),s.query(ActualWeeklyRevision).filter_by(company_id=owner['company_id']).count()); rows=s.query(ActualWeeklyObservation).filter_by(company_id=owner['company_id'],material_code='SKU-FU2-R2B',demand_type='sales').all();assert after==base and len(rows)==4 and len({r.period for r in rows})==4 and [float(r.quantity) for r in sorted(rows,key=lambda x:x.period)]==[101,102,103,104];assert s.query(RuntimeExecution).filter_by(company_id=owner['company_id']).count()==0 and s.query(ForecastVintage).filter_by(company_id=owner['company_id']).count()==0 and s.query(DecisionSnapshot).filter_by(company_id=owner['company_id']).count()==0;s.close();print(f'idempotency PASS after={after} company_scoped_hash=True',flush=True)
 finally:cleanup();print('FU2-R2B cleanup/residue PASS',flush=True)
if __name__=='__main__':main()
