"""FU2-R2C changed-value and Dataset-owned revision approval proof."""
import io,sys,hashlib
from pathlib import Path
from time import perf_counter
from uuid import uuid4
sys.path.insert(0,str(Path(__file__).resolve().parents[1]))
from openpyxl import load_workbook
from app.application.canonical_excel_ingestion import CanonicalExcelIngestionService,template_bytes
from app.application.actual_weekly_ledger import ActualWeeklyLedgerService
from app.auth import RegisterRequest,register
from app.database import SessionLocal
from app.models import ActualWeeklyObservation,ActualWeeklyRevision,DatasetVersion,RuntimeExecution,ForecastVintage,DecisionSnapshot
from scripts.verify_fu2_excel_ingestion import cleanup
def book(change=None):
 w=load_workbook(io.BytesIO(template_bytes()));s=w['Talep_Gecmisi'];s.delete_rows(2,s.max_row)
 for week,value in enumerate((101,102,103,104),1):s.append(['SKU-FU2-R2C','sales','finished_good',f'2026-W{week:02d}',change if week==3 and change is not None else value,'G','C'])
 o=io.BytesIO();w.save(o);return o.getvalue()
def main():
 cleanup();key=uuid4().hex;s=SessionLocal()
 try:o=register(RegisterRequest(email=f'fu2r2c-{key}@example.test',password='pilot-password-1',full_name='Pilot',company_name=f'FU2-R2C-{key}'),db=s)
 finally:s.close()
 try:
  service=CanonicalExcelIngestionService();base=book();changed=book(999);assert base!=changed;print('baseline_sha='+hashlib.sha256(base).hexdigest()+' changed_sha='+hashlib.sha256(changed).hexdigest(),flush=True)
  t=perf_counter();s=SessionLocal();d,_=service.stage(s,o['company_id'],o['user_id'],'base.xlsx',base);bid=d.id;s.close();print(f'baseline_stage={perf_counter()-t:.3f}s dataset={bid}',flush=True)
  t=perf_counter();s=SessionLocal();r=service.accept(s,o['company_id'],o['user_id'],bid);bvid=r['version_id'];s.close();print(f'baseline_accept={perf_counter()-t:.3f}s version={bvid}',flush=True)
  s=SessionLocal();a=s.query(ActualWeeklyObservation).filter_by(company_id=o['company_id'],material_code='SKU-FU2-R2C',period='2026-W03',demand_type='sales').one();assert float(a.quantity)==103;aid=a.id;s.close()
  t=perf_counter();s=SessionLocal();d,_=service.stage(s,o['company_id'],o['user_id'],'changed.xlsx',changed);cid=d.id;s.close();print(f'changed_stage={perf_counter()-t:.3f}s dataset={cid}',flush=True)
  s=SessionLocal();assert float(s.query(ActualWeeklyObservation).filter_by(id=aid).one().quantity)==103;assert s.query(ActualWeeklyRevision).filter_by(company_id=o['company_id'],source_dataset_id=cid).count()==0;s.close();print('pre_accept current=103 staged_ledger_revisions=0 (staging is Dataset-only)',flush=True)
  # Independent pending correction owned by baseline Dataset must survive changed acceptance.
  pending=ActualWeeklyLedgerService().ingest_dataset_actuals(o['company_id'],o['user_id'],bid,[{'material_code':'SKU-FU2-R2C','period':'2026-W04','quantity':777,'product_level':'finished_good','product_group':'G','product_class':'C'}],'sales');assert pending['proposed']==1
  s=SessionLocal();unrelated=s.query(ActualWeeklyRevision).filter_by(company_id=o['company_id'],source_dataset_id=bid,approval_status='proposed').one();unrelated_id=unrelated.id;s.close()
  t=perf_counter();s=SessionLocal();r=service.accept(s,o['company_id'],o['user_id'],cid);cvid=r['version_id'];s.close();print(f'changed_accept={perf_counter()-t:.3f}s version={cvid}',flush=True)
  s=SessionLocal();now=s.query(ActualWeeklyObservation).filter_by(id=aid).one();revs=s.query(ActualWeeklyRevision).filter_by(observation_id=aid).all();assert float(now.quantity)==999 and len(revs)==2;assert {(float(x.proposed_quantity),x.approval_status) for x in revs}=={(103.0,'accepted'),(999.0,'accepted')};assert s.query(ActualWeeklyRevision).filter_by(id=unrelated_id).one().approval_status=='proposed';vals={x.period:float(x.quantity) for x in s.query(ActualWeeklyObservation).filter_by(company_id=o['company_id'],material_code='SKU-FU2-R2C',demand_type='sales').all()};assert vals=={'2026-W01':101,'2026-W02':102,'2026-W03':999,'2026-W04':104};assert s.query(DatasetVersion).filter_by(dataset_id=bid).count()==1 and s.query(DatasetVersion).filter_by(dataset_id=cid).count()==1;assert s.query(RuntimeExecution).filter_by(company_id=o['company_id']).count()==0 and s.query(ForecastVintage).filter_by(company_id=o['company_id']).count()==0 and s.query(DecisionSnapshot).filter_by(company_id=o['company_id']).count()==0;s.close();print('fresh current=999 lineage=103 accepted -> 999 accepted; unrelated pending preserved; unchanged weeks PASS',flush=True)
 finally:cleanup();print('FU2-R2C cleanup/residue PASS',flush=True)
if __name__=='__main__':main()
