import io,json,sys,hashlib
from pathlib import Path
from time import perf_counter
from uuid import UUID,uuid4
sys.path.insert(0,str(Path(__file__).resolve().parents[1]))
from openpyxl import load_workbook
from app.application.canonical_excel_ingestion import CanonicalExcelIngestionService,template_bytes
from app.auth import RegisterRequest,register
from app.database import SessionLocal
from app.models import ActualWeeklyObservation as A,ActualWeeklyRevision as R,Dataset,DatasetVersion as V,DatasetValidationResult as VR,DatasetEvent as DE,Company,User,RuntimeExecution,ForecastVintage,DecisionSnapshot
from app.models.security import CompanyEncryptionKey
M=Path(__file__).with_suffix('.json'); CODE='SKU-FU2-R2C'
def wb(ch=None):
 w=load_workbook(io.BytesIO(template_bytes()));s=w['Talep_Gecmisi'];s.delete_rows(2,s.max_row)
 for k,v in enumerate((101,102,103,104),1):s.append([CODE,'sales','finished_good',f'2026-W{k:02d}',ch if k==3 and ch else v,'G','C'])
 o=io.BytesIO();w.save(o);return o.getvalue()
def data():return json.loads(M.read_text())
def setup():
 key=uuid4().hex;s=SessionLocal();o=register(RegisterRequest(email=f'r2cb-{key}@example.test',password='pilot-password-1',full_name='Pilot',company_name=f'FU2-R2CB-{key}'),db=s);s.close();x=CanonicalExcelIngestionService();b,c=wb(),wb(999);s=SessionLocal();d,_=x.stage(s,o['company_id'],o['user_id'],'b.xlsx',b);bid=str(d.id);s.close();s=SessionLocal();r=x.accept(s,o['company_id'],o['user_id'],bid);bvid=r['version_id'];s.close();s=SessionLocal();d,_=x.stage(s,o['company_id'],o['user_id'],'c.xlsx',c);cid=str(d.id);a=s.query(A).filter_by(company_id=o['company_id'],material_code=CODE,period='2026-W03',demand_type='sales').one();assert float(a.quantity)==103 and s.query(V).filter_by(dataset_id=cid).count()==0;s.close();M.write_text(json.dumps({**o,'baseline_dataset_id':bid,'baseline_version_id':bvid,'changed_dataset_id':cid,'actual_id':str(a.id),'company_name':f'FU2-R2CB-{key}','base_hash':hashlib.sha256(b).hexdigest(),'changed_hash':hashlib.sha256(c).hexdigest()}));print('R2C_B1_SETUP_COMPLETE',flush=True)
def accept():
 m=data();s=SessionLocal();a=s.query(A).filter_by(id=m['actual_id']).one();assert float(a.quantity)==103;t=perf_counter();r=CanonicalExcelIngestionService().accept(s,m['company_id'],m['user_id'],m['changed_dataset_id']);m['changed_version_id']=r['version_id'];M.write_text(json.dumps(m));s.close();s=SessionLocal();a=s.query(A).filter_by(id=m['actual_id']).one();assert float(a.quantity)==999 and s.query(A).filter_by(company_id=m['company_id'],material_code=CODE,period='2026-W03',demand_type='sales').count()==1;s.close();print(f'R2C_B2_CHANGED_ACCEPTANCE_COMPLETE seconds={perf_counter()-t:.3f}',flush=True)
def audit():
 m=data();s=SessionLocal();a=s.query(A).filter_by(id=m['actual_id']).one();rs=s.query(R).filter_by(observation_id=a.id).all();vals={x.period:float(x.quantity) for x in s.query(A).filter_by(company_id=m['company_id'],material_code=CODE,demand_type='sales').all()};assert float(a.quantity)==999 and {(float(x.proposed_quantity),x.approval_status) for x in rs}>={(103.,'accepted'),(999.,'accepted')} and vals=={'2026-W01':101.,'2026-W02':102.,'2026-W03':999.,'2026-W04':104.};assert s.query(V).filter_by(id=m['baseline_version_id']).count()==1 and s.query(V).filter_by(id=m['changed_version_id']).count()==1 and m['baseline_version_id']!=m['changed_version_id'];assert s.query(RuntimeExecution).filter_by(company_id=m['company_id']).count()==0 and s.query(ForecastVintage).filter_by(company_id=m['company_id']).count()==0 and s.query(DecisionSnapshot).filter_by(company_id=m['company_id']).count()==0;print('model=one stable Actual row plus immutable revisions; lineage='+str([(float(x.proposed_quantity),x.approval_status,str(x.source_dataset_id)) for x in rs]),flush=True);s.close();print('R2C_B3_AUDIT_COMPLETE',flush=True)
def cleanup():
 m=data();s=SessionLocal();cid=UUID(m['company_id']);ds=[UUID(m['baseline_dataset_id']),UUID(m['changed_dataset_id'])];s.query(R).filter_by(company_id=cid).delete(synchronize_session=False);s.query(A).filter_by(company_id=cid).delete(synchronize_session=False);s.query(DE).filter(DE.dataset_id.in_(ds)).delete(synchronize_session=False);s.query(VR).filter(VR.dataset_id.in_(ds)).delete(synchronize_session=False);s.query(V).filter(V.dataset_id.in_(ds)).delete(synchronize_session=False);s.query(Dataset).filter(Dataset.id.in_(ds)).delete(synchronize_session=False);s.query(CompanyEncryptionKey).filter_by(company_id=cid).delete(synchronize_session=False);s.query(User).filter_by(company_id=cid).delete(synchronize_session=False);s.query(Company).filter_by(id=cid).delete(synchronize_session=False);s.commit();assert s.query(Company).filter_by(id=cid).count()==0;s.close();M.unlink();print('R2C_B4_CLEANUP_COMPLETE',flush=True)
if __name__=='__main__':{'setup':setup,'accept':accept,'audit':audit,'cleanup':cleanup}[sys.argv[1]]()
