"""Bounded persisted acceptance proof for the FU2 R2E identity matrix."""

import io
import sys
from pathlib import Path
from time import perf_counter
from uuid import UUID, uuid4

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.application.canonical_excel_ingestion import (
    CanonicalExcelIngestionService,
    template_bytes,
)
from app.auth import RegisterRequest, register
from app.database import SessionLocal
from app.models import (
    ActualWeeklyObservation as Actual,
    ActualWeeklyRevision as ActualRevision,
    Company,
    Dataset,
    DatasetEvent,
    DatasetValidationResult,
    DatasetVersion,
    DecisionSnapshot,
    ForecastVintage,
    RuntimeExecution,
    User,
)
from app.models.security import CompanyEncryptionKey


PERIODS_AND_QUANTITIES = {
    "2026-W01": 101.0,
    "2026-W02": 102.0,
    "2026-W03": 103.0,
    "2026-W04": 104.0,
}


def workbook_bytes(material_code: str, demand_type: str, product_level: str) -> bytes:
    workbook = load_workbook(io.BytesIO(template_bytes()))
    sheet = workbook["Talep_Gecmisi"]
    sheet.delete_rows(2, sheet.max_row)
    for period, quantity in PERIODS_AND_QUANTITIES.items():
        sheet.append(
            [material_code, demand_type, product_level, period, quantity, "FU2-G", "FU2-C"]
        )
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def cleanup(company_id: str) -> None:
    """Delete only the explicitly-owned, dependency-ordered synthetic graph."""
    company_uuid = UUID(company_id)
    session = SessionLocal()
    try:
        dataset_ids = [
            dataset.id for dataset in session.query(Dataset).filter_by(company_id=company_uuid).all()
        ]
        session.query(ActualRevision).filter_by(company_id=company_uuid).delete(
            synchronize_session=False
        )
        session.query(Actual).filter_by(company_id=company_uuid).delete(synchronize_session=False)
        if dataset_ids:
            session.query(DatasetEvent).filter(DatasetEvent.dataset_id.in_(dataset_ids)).delete(
                synchronize_session=False
            )
            session.query(DatasetValidationResult).filter(
                DatasetValidationResult.dataset_id.in_(dataset_ids)
            ).delete(synchronize_session=False)
            session.query(DatasetVersion).filter(DatasetVersion.dataset_id.in_(dataset_ids)).delete(
                synchronize_session=False
            )
            session.query(Dataset).filter(Dataset.id.in_(dataset_ids)).delete(synchronize_session=False)
        session.query(CompanyEncryptionKey).filter_by(company_id=company_uuid).delete(
            synchronize_session=False
        )
        session.query(User).filter_by(company_id=company_uuid).delete(synchronize_session=False)
        session.query(Company).filter_by(id=company_uuid).delete(synchronize_session=False)
        session.commit()
        assert session.query(Company).filter_by(id=company_uuid).count() == 0
        assert session.query(Actual).filter_by(company_id=company_uuid).count() == 0
        assert session.query(Dataset).filter_by(company_id=company_uuid).count() == 0
    finally:
        session.close()


def assert_fresh_persistence(
    company_id: str, material_code: str, demand_type: str, product_level: str, dataset_id: str
) -> tuple[list[str], float]:
    """Use a new session and primitive identifiers to reconstruct accepted evidence."""
    started = perf_counter()
    session = SessionLocal()
    try:
        company_uuid = UUID(company_id)
        rows = (
            session.query(Actual)
            .filter_by(
                company_id=company_uuid,
                material_code=material_code,
                demand_type=demand_type,
            )
            .order_by(Actual.period)
            .all()
        )
        actual_values = {row.period: float(row.quantity) for row in rows}
        assert actual_values == PERIODS_AND_QUANTITIES
        assert all(
            row.product_level == product_level
            and row.product_group == "FU2-G"
            and row.product_class == "FU2-C"
            for row in rows
        )
        assert session.query(DatasetVersion).filter_by(dataset_id=UUID(dataset_id)).count() == 1
        assert session.query(RuntimeExecution).filter_by(company_id=company_uuid).count() == 0
        assert session.query(ForecastVintage).filter_by(company_id=company_uuid).count() == 0
        assert session.query(DecisionSnapshot).filter_by(company_id=company_uuid).count() == 0
        return [str(row.id) for row in rows], perf_counter() - started
    finally:
        session.close()


def run_scenario(product_level: str, demand_type: str, material_code: str) -> None:
    token = uuid4().hex
    company_id = None
    try:
        session = SessionLocal()
        owner = register(
            RegisterRequest(
                email=f"fu2-r2e-{token}@example.test",
                password="pilot-password-1",
                full_name="FU2 R2E Owner",
                company_name=f"FU2 R2E {token}",
            ),
            db=session,
        )
        company_id = owner["company_id"]
        session.close()

        service = CanonicalExcelIngestionService()
        started = perf_counter()
        session = SessionLocal()
        dataset, reused = service.stage(
            session,
            company_id,
            owner["user_id"],
            f"{material_code}.xlsx",
            workbook_bytes(material_code, demand_type, product_level),
        )
        dataset_id = str(dataset.id)
        assert reused is False
        assert dataset.validations[-1].is_valid is True
        session.close()
        stage_seconds = perf_counter() - started

        started = perf_counter()
        session = SessionLocal()
        accepted = service.accept(session, company_id, owner["user_id"], dataset_id)
        session.close()
        accept_seconds = perf_counter() - started
        assert accepted["status"] == "READY_FOR_WORKFLOW"

        actual_ids, fresh_read_seconds = assert_fresh_persistence(
            company_id, material_code, demand_type, product_level, dataset_id
        )
        print(
            "R2E PASS "
            f"scope={product_level}/{demand_type} material={material_code} "
            f"dataset={dataset_id} version={accepted['version_id']} actual_ids={actual_ids} "
            f"stage={stage_seconds:.3f}s accept={accept_seconds:.3f}s fresh_read={fresh_read_seconds:.3f}s",
            flush=True,
        )
    finally:
        if company_id is not None:
            cleanup(company_id)
            print(f"R2E CLEANUP PASS scope={product_level}/{demand_type}", flush=True)


if __name__ == "__main__":
    run_scenario(*sys.argv[1:])
