"""Resumable mounted-API FU2 R2F tenant-isolation proof."""
import hashlib, io, json, sys
from pathlib import Path
from time import perf_counter
from uuid import UUID, uuid4
from fastapi.testclient import TestClient
from openpyxl import load_workbook
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app.application.canonical_excel_ingestion import template_bytes
from app.database import SessionLocal
from app.main import app
from app.models import ActualWeeklyObservation as A, ActualWeeklyRevision as R, Company, Dataset, DatasetEvent as DE, DatasetValidationResult as VR, DatasetVersion as V, RuntimeExecution, ForecastVintage, DecisionSnapshot, User, WorkflowExecution
from app.models.security import CompanyEncryptionKey
M=Path(__file__).with_suffix('.json'); W=Path(__file__).with_suffix('.xlsx'); CODE='SKU-FU2-R2F'; VALS={f'2026-W{x:02d}':float(100+x) for x in range(1,5)}
def data(): return json.loads(M.read_text())
def headers(c,o):
 r=c.post('/auth/login',json={'email':o['email'],'password':'pilot-password-1'});assert r.status_code==200,r.text;return {'Authorization':'Bearer '+r.json()['access_token']}
def wb():
 w=load_workbook(io.BytesIO(template_bytes()));s=w['Talep_Gecmisi'];s.delete_rows(2,s.max_row)
 for p,q in VALS.items():s.append([CODE,'sales','finished_good',p,q,'FU2-G','FU2-C'])
 b=io.BytesIO();w.save(b);W.write_bytes(b.getvalue());return b.getvalue()
def upload(c,h):
 r=c.post('/api/v2/dataset/pilot/upload',headers=h,files={'file':('same.xlsx',W.read_bytes(),'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')});assert r.status_code==200,r.text;d=r.json();assert d['READY_FOR_ACCEPTANCE'];return d
def counts(cid):
 s=SessionLocal();u=UUID(cid)
 try:return {n:s.query(x).filter_by(company_id=u).count() for n,x in {'runtime':RuntimeExecution,'vintage':ForecastVintage,'decision':DecisionSnapshot}.items()}|{'workflow':s.query(WorkflowExecution).filter_by(company_id=u).count() if hasattr(WorkflowExecution,'company_id') else 0}
 finally:s.close()
def clean(m):
 s=SessionLocal()
 try:
  cs=[UUID(o['company_id']) for o in m['owners']];ds=[UUID(x) for x in m.get('datasets',[]) if x]
  s.query(R).filter(R.company_id.in_(cs)).delete(synchronize_session=False);s.query(A).filter(A.company_id.in_(cs)).delete(synchronize_session=False);s.query(DE).filter(DE.dataset_id.in_(ds)).delete(synchronize_session=False);s.query(VR).filter(VR.dataset_id.in_(ds)).delete(synchronize_session=False);s.query(V).filter(V.dataset_id.in_(ds)).delete(synchronize_session=False);s.query(Dataset).filter(Dataset.id.in_(ds)).delete(synchronize_session=False);s.query(CompanyEncryptionKey).filter(CompanyEncryptionKey.company_id.in_(cs)).delete(synchronize_session=False);s.query(User).filter(User.company_id.in_(cs)).delete(synchronize_session=False);s.query(Company).filter(Company.id.in_(cs)).delete(synchronize_session=False)
  s.commit()
 finally:s.close()
def setup():
 m={'owners':[],'datasets':[None,None],'versions':[None,None],'expected':VALS};wb()
 with TestClient(app) as c:
  for tag in 'AB':
   e=f'fu2-r2f-{tag}-{uuid4().hex[:10]}@example.test';r=c.post('/auth/register',json={'email':e,'password':'pilot-password-1','full_name':tag,'company_name':f'FU2-R2F-{tag}-{uuid4().hex[:8]}'});assert r.status_code==201,r.text;o=r.json();o['email']=e;m['owners'].append(o);headers(c,o)
 m['sha256']=hashlib.sha256(W.read_bytes()).hexdigest();M.write_text(json.dumps(m));print('FU2_R2F_SETUP_COMPLETE')
def stage(i):
 m=data();t=perf_counter()
 with TestClient(app) as c:d=upload(c,headers(c,m['owners'][i]))
 m['datasets'][i]=d['dataset_id'];s=SessionLocal();row=s.query(Dataset).filter_by(id=UUID(d['dataset_id']),company_id=UUID(m['owners'][i]['company_id'])).one();assert row.validations[-1].is_valid;fp=row.dataset_hash
 if i==1:
  a=s.query(Dataset).filter_by(id=UUID(m['datasets'][0]),company_id=UUID(m['owners'][0]['company_id'])).one();assert d['dataset_id']!=str(a.id) and fp!=m['fingerprints'][0];assert s.query(V).filter_by(dataset_id=row.id).count()==0 and s.query(R).filter_by(company_id=row.company_id,source_dataset_id=row.id).count()==0
 s.close();m.setdefault('fingerprints',[None,None])[i]=fp;m.setdefault('upload_seconds',[None,None])[i]=perf_counter()-t;M.write_text(json.dumps(m));print(f'Dataset {"AB"[i]} ID {d["dataset_id"]} fingerprint={fp} seconds={m["upload_seconds"][i]:.3f}');print(f'FU2_R2F_UPLOAD_{"AB"[i]}_COMPLETE')
def isolate():
 m=data()
 with TestClient(app) as c:
  ha,hb=headers(c,m['owners'][0]),headers(c,m['owners'][1]);a,b=upload(c,ha),upload(c,hb);assert a['same_file_retry'] and a['dataset_id']==m['datasets'][0] and b['same_file_retry'] and b['dataset_id']==m['datasets'][1]
  for h,d in ((ha,m['datasets'][1]),(hb,m['datasets'][0])):
   r=c.post(f'/api/v2/dataset/pilot/{d}/accept',headers=h);assert r.status_code==404 and r.json()=={'detail':'DATASET_UNAVAILABLE'}
 s=SessionLocal();assert all(s.query(V).filter_by(dataset_id=UUID(d)).count()==0 for d in m['datasets']);assert s.query(A).filter(A.company_id.in_([UUID(x['company_id']) for x in m['owners']])).count()==0;s.close();print('FU2_R2F_ISOLATION_COMPLETE')
def accept(i):
 m=data();t=perf_counter()
 with TestClient(app) as c:r=c.post(f'/api/v2/dataset/pilot/{m["datasets"][i]}/accept',headers=headers(c,m['owners'][i]));assert r.status_code==200,r.text;d=r.json();assert d['status']=='READY_FOR_WORKFLOW';m['versions'][i]=d['version_id']
 M.write_text(json.dumps(m));assert counts(m['owners'][i]['company_id'])=={'runtime':0,'vintage':0,'decision':0,'workflow':0};print(f'FU2_R2F_ACCEPT_{"AB"[i]}_COMPLETE seconds={perf_counter()-t:.3f}')
def audit():
 m=data();s=SessionLocal()
 try:
  for i,o in enumerate(m['owners']):
   c=UUID(o['company_id']);d=UUID(m['datasets'][i]);assert s.query(V).filter_by(id=UUID(m['versions'][i]),dataset_id=d).count()==1;rows=s.query(A).filter_by(company_id=c,material_code=CODE,demand_type='sales').all();assert {x.period:float(x.quantity) for x in rows}==VALS and all(x.product_level=='finished_good' for x in rows);assert s.query(R).filter_by(company_id=c,source_dataset_id=d,approval_status='accepted').count()==4;assert counts(o['company_id'])=={'runtime':0,'vintage':0,'decision':0,'workflow':0}
  assert s.query(CompanyEncryptionKey).filter_by(company_id=UUID(m['owners'][0]['company_id'])).one().id != s.query(CompanyEncryptionKey).filter_by(company_id=UUID(m['owners'][1]['company_id'])).one().id
 finally:s.close()
 print('FU2_R2F_FINAL_AUDIT_COMPLETE')
def audit_a():
 m=data();s=SessionLocal()
 try:
  ca,cb=UUID(m['owners'][0]['company_id']),UUID(m['owners'][1]['company_id']);da,db=UUID(m['datasets'][0]),UUID(m['datasets'][1]);va=UUID(m['versions'][0]);a=s.query(Dataset).filter_by(id=da,company_id=ca).one();b=s.query(Dataset).filter_by(id=db,company_id=cb).one();assert a.state.value=='approved' and b.state.value=='validated';assert s.query(V).filter_by(dataset_id=da).count()==1 and s.query(V).filter_by(id=va,dataset_id=da).count()==1 and s.query(V).filter_by(dataset_id=db).count()==0
  rows=s.query(A).filter_by(company_id=ca,material_code=CODE,demand_type='sales').all();assert {x.period:float(x.quantity) for x in rows}==VALS and all(x.product_level=='finished_good' and x.product_group=='FU2-G' and x.product_class=='FU2-C' for x in rows);assert len({(x.company_id,x.material_code,x.demand_type,x.period) for x in rows})==4
  rev=s.query(R).filter_by(company_id=ca,source_dataset_id=da,approval_status='accepted').all();assert len(rev)==4 and s.query(R).filter_by(company_id=cb,source_dataset_id=db).count()==0 and s.query(A).filter_by(company_id=cb,material_code=CODE,demand_type='sales').count()==0;assert counts(str(ca))=={'runtime':0,'vintage':0,'decision':0,'workflow':0} and counts(str(cb))=={'runtime':0,'vintage':0,'decision':0,'workflow':0}
  m['dataset_a_audit']={'actual_count':len(rows),'accepted_revision_count':len(rev),'status':'READY_FOR_WORKFLOW'};M.write_text(json.dumps(m))
 finally:s.close()
 print('FU2_R2F_ACCEPT_A_COMPLETE')
def cleanup():
 m=data();clean(m);s=SessionLocal();assert all(s.query(Company).filter_by(id=UUID(o['company_id'])).count()==0 for o in m['owners']);s.close();M.unlink();W.unlink(missing_ok=True);print('FU2_R2F_CLEANUP_COMPLETE')
if __name__=='__main__':{'setup':setup,'upload_a':lambda:stage(0),'upload_b':lambda:stage(1),'isolate':isolate,'accept_a':lambda:accept(0),'accept_b':lambda:accept(1),'audit_a':audit_a,'audit':audit,'cleanup':cleanup}[sys.argv[1]]()
