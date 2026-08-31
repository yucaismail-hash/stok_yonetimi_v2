"""Disposable PostgreSQL proof for the Official Excel V3 acceptance boundary."""
import io
import os
import sys
from pathlib import Path
from uuid import uuid4

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook
from sqlalchemy import create_engine, inspect
from sqlalchemy.orm import sessionmaker

from app.application.actual_weekly_ledger import ActualWeeklyLedgerService
from app.application.business_workflow_acceptance import BusinessWorkflowAcceptanceService
from app.application.business_workflow_readiness import BusinessWorkflowReadinessService
from app.application.canonical_excel_ingestion import CanonicalExcelIngestionService, template_bytes
from app.engine.capability_contracts import CapabilityExecutionRequest
from app.engine.capability_registry import Capability
from app.engine.dataset_runtime_provider import DatasetRuntimeProvider
from app.models import ActualWeeklyObservation, ActualWeeklyRevision, Company, CompanyEncryptionKey, Dataset, DatasetEvent, DatasetValidationResult, DatasetVersion, DatasetVersionProductInput, RuntimeExecution, RuntimeTask, User


URL = os.environ["DATABASE_URL"]
ENGINE = create_engine(URL)
Session = sessionmaker(bind=ENGINE, autoflush=False, autocommit=False)


def workbook(*, week_three=None, blank_economics=False, through_week=52):
    book = load_workbook(io.BytesIO(template_bytes()))
    sheet = book["Temel_Veriler"]
    headers = [cell.value for cell in sheet[1]]
    if week_three is not None:
        sheet.cell(2, headers.index("2026-W03") + 1).value = week_three
    if blank_economics:
        for field in ("Birim Maliyet (TL)", "Stok Tutma Oranı (%)", "Stok Tükenme Maliyeti"):
            sheet.cell(2, headers.index(field) + 1).value = None
        for week in range(through_week + 1, 53):
            sheet.cell(2, headers.index(f"2026-W{week:02d}") + 1).value = None
    data = io.BytesIO(); book.save(data); return data.getvalue()


def incremental_workbook(*, new_period, new_value, corrected_week_three=None):
    """A one-week V3 upload, optionally with one historical correction."""
    book = load_workbook(io.BytesIO(template_bytes()))
    sheet = book["Temel_Veriler"]
    headers = [cell.value for cell in sheet[1]]
    for header in headers:
        if isinstance(header, str) and header.startswith("2026-W"):
            sheet.cell(2, headers.index(header) + 1).value = None
    next_column = len(headers) + 1
    sheet.cell(1, next_column).value = new_period
    sheet.cell(2, next_column).value = new_value
    if corrected_week_three is not None:
        sheet.cell(2, headers.index("2026-W03") + 1).value = corrected_week_three
    data = io.BytesIO(); book.save(data); return data.getvalue()


def invalid_supplier_reference_workbook():
    book = load_workbook(io.BytesIO(template_bytes()))
    book["Malzeme_Tedarikciler"].append(["MISSING-SKU", "MISSING-SUPPLIER", 0.5, None, None])
    data = io.BytesIO(); book.save(data); return data.getvalue()


def owner(session, label):
    company = Company(name=f"FU2V3A-{label}-{uuid4().hex}", tax_id=uuid4().hex)
    session.add(company); session.flush()
    user = User(company_id=company.id, email=f"{uuid4().hex}@fu2-v3a.invalid", hashed_password="test")
    session.add(user); session.commit()
    return company, user


def request(company, user, dataset, demand, capability=Capability.DEMAND_FORECAST):
    return CapabilityExecutionRequest(uuid4(), "v3", "forecast", capability, company.id, user.id, dataset.id, 30, material_codes=["SKU-001"], params={"demand_type": demand, "forecast_cutoff_period": "2026-W52"})


def cleanup(session, company_ids):
    datasets = [row.id for row in session.query(Dataset).filter(Dataset.company_id.in_(company_ids)).all()]
    versions = [row.id for row in session.query(DatasetVersion).filter(DatasetVersion.dataset_id.in_(datasets)).all()]
    executions = [row.execution_id for row in session.query(RuntimeExecution).filter(RuntimeExecution.company_id.in_(company_ids)).all()]
    session.query(RuntimeTask).filter(RuntimeTask.execution_id.in_(executions)).delete(synchronize_session=False)
    session.query(RuntimeExecution).filter(RuntimeExecution.execution_id.in_(executions)).delete(synchronize_session=False)
    session.query(ActualWeeklyRevision).filter(ActualWeeklyRevision.company_id.in_(company_ids)).delete(synchronize_session=False)
    session.query(ActualWeeklyObservation).filter(ActualWeeklyObservation.company_id.in_(company_ids)).delete(synchronize_session=False)
    session.query(DatasetVersionProductInput).filter(DatasetVersionProductInput.dataset_version_id.in_(versions)).delete(synchronize_session=False)
    session.query(DatasetEvent).filter(DatasetEvent.dataset_id.in_(datasets)).delete(synchronize_session=False)
    session.query(DatasetValidationResult).filter(DatasetValidationResult.dataset_id.in_(datasets)).delete(synchronize_session=False)
    session.query(DatasetVersion).filter(DatasetVersion.id.in_(versions)).delete(synchronize_session=False)
    session.query(Dataset).filter(Dataset.id.in_(datasets)).delete(synchronize_session=False)
    session.query(CompanyEncryptionKey).filter(CompanyEncryptionKey.company_id.in_(company_ids)).delete(synchronize_session=False)
    session.query(User).filter(User.company_id.in_(company_ids)).delete(synchronize_session=False)
    session.query(Company).filter(Company.id.in_(company_ids)).delete(synchronize_session=False)
    session.commit()


def main():
    session = Session(); companies = []
    try:
        nullable = {column["name"]: column["nullable"] for column in inspect(ENGINE).get_columns("dataset_version_product_inputs")}
        assert nullable["unit_cost"] and nullable["holding_rate"] and nullable["stockout_cost"]
        print("OPTIONAL_ECONOMICS_SCHEMA_PASS")
        service = CanonicalExcelIngestionService(); company, user = owner(session, "sales"); companies.append(company.id)
        # Forced post-product-input failure: the shared acceptance transaction must roll back all writes.
        failed, _ = service.stage(session, company.id, user.id, "failed.xlsx", workbook(), demand_type="sales")
        original = ActualWeeklyLedgerService.ingest_dataset_actuals_in_session
        def fail_after_inputs(*args, **kwargs): raise RuntimeError("forced acceptance failure")
        ActualWeeklyLedgerService.ingest_dataset_actuals_in_session = fail_after_inputs
        try:
            try: service.accept(session, company.id, user.id, failed.id)
            except RuntimeError: pass
            else: raise AssertionError("forced failure was not raised")
        finally:
            ActualWeeklyLedgerService.ingest_dataset_actuals_in_session = original
        session.expire_all()
        assert session.query(DatasetVersion).filter_by(dataset_id=failed.id).count() == 0
        assert session.query(DatasetVersionProductInput).filter_by(company_id=company.id).count() == 0
        assert session.query(ActualWeeklyRevision).filter_by(source_dataset_id=failed.id).count() == 0
        assert session.query(Dataset).filter_by(id=failed.id).one().state.value == "validated"
        print("ATOMIC_ROLLBACK_PASS")

        dataset, retry = service.stage(session, company.id, user.id, "sales.xlsx", workbook(), demand_type="sales", service_level={"mode":"manual", "value":0.91})
        assert not retry; accepted = service.accept(session, company.id, user.id, dataset.id); version_id = accepted["version_id"]
        session.expire_all()
        assert session.query(DatasetVersionProductInput).filter_by(dataset_version_id=version_id).count() == 1
        assert session.query(ActualWeeklyObservation).filter_by(company_id=company.id, demand_type="sales").count() == 52
        assert session.query(ActualWeeklyRevision).filter_by(company_id=company.id, approval_status="accepted").count() == 52
        prepared = DatasetRuntimeProvider(session)(request(company, user, dataset, "sales"))
        item = prepared["items"][0]
        assert len(item["demand_history"]) == 52
        safety_item = DatasetRuntimeProvider(session)(request(company, user, dataset, "sales", Capability.SAFETY_STOCK))["items"][0]
        simulation_item = DatasetRuntimeProvider(session)(request(company, user, dataset, "sales", Capability.SIMULATION))["items"][0]
        backtest_item = DatasetRuntimeProvider(session)(request(company, user, dataset, "sales", Capability.BACKTEST))["items"][0]
        assert safety_item["lead_time_days"] == 7.0 and simulation_item["initial_stock"] == 250.0 and simulation_item["eoq"] == 50.0
        assert len(backtest_item["demand_history"]) == 52 and backtest_item["lead_time_days"] == 7.0
        assert prepared["service_level"] == {"mode":"manual", "value":0.91}
        assert service.accept(session, company.id, user.id, dataset.id)["idempotent"]
        print("ACCEPT_RUNTIME_IDEMPOTENCY_PASS")

        # The later upload contains one period only.  Readiness must resolve the
        # persisted effective history, rather than the latest workbook payload.
        incremental, _ = service.stage(session, company.id, user.id, "incremental.xlsx", incremental_workbook(new_period="2027-W01", new_value=53), demand_type="sales")
        service.accept(session, company.id, user.id, incremental.id); session.expire_all()
        effective = DatasetRuntimeProvider(session).preflight(request(company, user, incremental, "sales"))["items"][0]
        readiness = BusinessWorkflowReadinessService().evaluate(session, company.id, user.id, incremental.id)
        assert len(effective["demand_history"]) == 53 and effective["history_periods"][-1] == "2027-W01"
        assert readiness.materials[0].available_weeks == 53 and readiness.status == "READY"
        print("EFFECTIVE_HISTORY_52_PLUS_1_PASS")

        correction_company, correction_user = owner(session, "correction"); companies.append(correction_company.id)
        correction_base, _ = service.stage(session, correction_company.id, correction_user.id, "correction-base.xlsx", workbook(), demand_type="sales")
        service.accept(session, correction_company.id, correction_user.id, correction_base.id)
        correction_delta, _ = service.stage(session, correction_company.id, correction_user.id, "correction-delta.xlsx", incremental_workbook(new_period="2027-W01", new_value=53, corrected_week_three=999), demand_type="sales")
        service.accept(session, correction_company.id, correction_user.id, correction_delta.id); session.expire_all()
        corrected_item = DatasetRuntimeProvider(session).preflight(request(correction_company, correction_user, correction_delta, "sales"))["items"][0]
        corrected_readiness = BusinessWorkflowReadinessService().evaluate(session, correction_company.id, correction_user.id, correction_delta.id)
        assert len(corrected_item["demand_history"]) == 53 and corrected_item["demand_history"][2] == 999.0
        assert corrected_readiness.materials[0].available_weeks == 53
        print("EFFECTIVE_HISTORY_CORRECTION_NO_DOUBLE_COUNT_PASS")

        optional, _ = service.stage(session, company.id, user.id, "optional-economics.xlsx", workbook(blank_economics=True, through_week=12), demand_type="sales")
        validation = session.query(DatasetValidationResult).filter_by(dataset_id=optional.id).one()
        assert validation.is_valid and len(validation.errors or []) == 0 and len(validation.warnings or []) == 3
        optional_accept = service.accept(session, company.id, user.id, optional.id)
        optional_input = session.query(DatasetVersionProductInput).filter_by(dataset_version_id=optional_accept["version_id"]).one()
        assert optional_input.unit_cost is None and optional_input.holding_rate is None and optional_input.stockout_cost is None
        internal_item = DatasetRuntimeProvider(session)._official_v3_items(optional, request(company, user, optional, "sales"), {"demand_type": "sales"})[0]
        assert internal_item["unit_cost"] is None and internal_item["holding_rate"] is None and internal_item["stockout_cost"] is None
        print("OPTIONAL_ECONOMICS_NULL_PERSISTENCE_PASS")

        invalid_reference, _ = service.stage(session, company.id, user.id, "invalid-reference.xlsx", invalid_supplier_reference_workbook(), demand_type="sales")
        invalid_validation = session.query(DatasetValidationResult).filter_by(dataset_id=invalid_reference.id).one()
        issue = next(row for row in invalid_validation.errors if row["column"] == "Ürün Kodu")
        assert not invalid_validation.is_valid and issue["severity"] == "ERROR" and "MISSING-SKU" in issue["message"]
        try:
            service.accept(session, company.id, user.id, invalid_reference.id)
        except Exception as exc:
            assert str(exc) == "DATASET_NOT_READY_FOR_ACCEPTANCE"
        else:
            raise AssertionError("invalid cross-reference acceptance was not blocked")
        print("CROSS_REFERENCE_BLOCKING_PASS")

        # Acceptance plans the analytical graph only; it executes no capability here.
        import app.application.business_workflow_acceptance as workflow_module
        from app.application.forecast_scope import ForecastScopeService
        original_scope = workflow_module.ForecastScopeService
        workflow_module.ForecastScopeService = lambda: ForecastScopeService(Session)
        try:
            execution = BusinessWorkflowAcceptanceService(session_factory=Session).accept_or_resolve(company.id, user.id, dataset.id)
        finally:
            workflow_module.ForecastScopeService = original_scope
        tasks = session.query(RuntimeTask).filter_by(execution_id=execution.execution_id, company_id=company.id).order_by(RuntimeTask.task_order).all()
        assert [task.task_id for task in tasks] == ["forecast", "safety_stock", "simulation", "backtest"]
        assert all(task.capability != "decision" for task in tasks)
        print("BUSINESS_WORKFLOW_INPUT_AND_BACKTEST_READY_PASS")

        corrected, _ = service.stage(session, company.id, user.id, "corrected.xlsx", workbook(week_three=999), demand_type="sales")
        service.accept(session, company.id, user.id, corrected.id); session.expire_all()
        old_input = session.query(DatasetVersionProductInput).filter_by(dataset_version_id=version_id).one()
        assert float(old_input.initial_stock) == 250.0
        assert session.query(ActualWeeklyObservation).filter_by(company_id=company.id, period="2026-W03", demand_type="sales").one().quantity == 999
        print("CORRECTION_VERSION_PRESERVATION_PASS")

        other, other_user = owner(session, "consumption"); companies.append(other.id)
        consumption, _ = service.stage(session, other.id, other_user.id, "consumption.xlsx", workbook(), demand_type="consumption")
        service.accept(session, other.id, other_user.id, consumption.id); session.expire_all()
        consumption_item = DatasetRuntimeProvider(session)(request(other, other_user, consumption, "consumption"))["items"][0]
        assert consumption_item["demand_history"][0] == 40.0
        try: DatasetRuntimeProvider(session)(request(other, other_user, dataset, "sales"))
        except Exception: pass
        else: raise AssertionError("cross-tenant dataset was resolved")
        print("DEMAND_AND_TENANT_ISOLATION_PASS")
    finally:
        cleanup(session, companies); session.close()
    print("FU2_V3A_POSTGRES_PASS")


if __name__ == "__main__": main()
