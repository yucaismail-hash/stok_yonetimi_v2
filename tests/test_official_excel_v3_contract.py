"""Focused, database-free Official Excel V3 contract tests."""
import io
import unittest

from openpyxl import load_workbook

from app.application.canonical_excel_ingestion import (
    OFFICIAL_V3_HEADERS,
    OFFICIAL_V3_SHEET,
    parse_workbook,
    parse_workbook_details,
    template_bytes,
)


class OfficialExcelV3ContractTests(unittest.TestCase):
    @staticmethod
    def workbook(change=None):
        book = load_workbook(io.BytesIO(template_bytes()))
        if change:
            change(book)
        output = io.BytesIO(); book.save(output)
        return output.getvalue()

    def test_template_has_official_sheets_and_wide_iso_headers(self):
        book = load_workbook(io.BytesIO(template_bytes()), read_only=True, data_only=True)
        self.assertEqual(book.sheetnames, [OFFICIAL_V3_SHEET, "Malzeme_Tedarikciler", "Tedarikciler", "Events", "Data_Requirement_Matrix"])
        headers = list(next(book[OFFICIAL_V3_SHEET].iter_rows(max_row=1, values_only=True)))
        self.assertEqual(tuple(headers[:len(OFFICIAL_V3_HEADERS)]), OFFICIAL_V3_HEADERS)
        self.assertEqual(headers[len(OFFICIAL_V3_HEADERS):], [f"2026-W{week:02d}" for week in range(1, 53)])

    def test_template_normalizes_wide_rows_to_weekly_canonical_rows(self):
        rows, errors = parse_workbook(template_bytes())
        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 52)
        self.assertEqual(rows[0]["material_code"], "SKU-001")
        self.assertEqual(rows[0]["product_level"], "finished_good")
        self.assertEqual(rows[0]["period"], "2026-W01")
        self.assertEqual(rows[-1]["period"], "2026-W52")
        self.assertNotIn("demand_type", rows[0])

    def test_optional_economics_are_warnings_not_errors_and_future_weeks_are_legal(self):
        def change(book):
            sheet = book[OFFICIAL_V3_SHEET]
            headers = [cell.value for cell in sheet[1]]
            for field in ("Birim Maliyet (TL)", "Stok Tutma Oranı (%)", "Stok Tükenme Maliyeti"):
                sheet.cell(2, headers.index(field) + 1).value = None
            for week in range(13, 53): sheet.cell(2, headers.index(f"2026-W{week:02d}") + 1).value = None
        rows, errors, warnings, _ = parse_workbook_details(self.workbook(change))
        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 12)
        self.assertEqual({row["unit_cost"] for row in rows}, {None})
        self.assertEqual({warning["column"] for warning in warnings}, {"Birim Maliyet (TL)", "Stok Tutma Oranı (%)", "Stok Tükenme Maliyeti"})
        self.assertTrue(all(warning["severity"] == "WARNING" and warning["row"] is None for warning in warnings))

    def test_required_and_numeric_errors_identify_exact_column(self):
        def change(book):
            sheet = book[OFFICIAL_V3_SHEET]; headers = [cell.value for cell in sheet[1]]
            sheet.cell(2, headers.index("Ürün Grubu") + 1).value = None
            sheet.cell(2, headers.index("Tedarik Süresi (Gün)") + 1).value = 0
        _, errors, _, _ = parse_workbook_details(self.workbook(change))
        self.assertIn(("REQUIRED_VALUE_MISSING", "Ürün Grubu"), {(error["code"], error["column"]) for error in errors})
        self.assertIn(("VALUE_OUT_OF_RANGE", "Tedarik Süresi (Gün)"), {(error["code"], error["column"]) for error in errors})
        self.assertTrue(all(error["sheet"] == OFFICIAL_V3_SHEET and error["row"] == 2 for error in errors))

    def test_invalid_product_level_identifies_column_and_allowed_values(self):
        def change(book): book[OFFICIAL_V3_SHEET].cell(2, 5).value = "Bilinmeyen"
        _, errors, _, _ = parse_workbook_details(self.workbook(change))
        issue = next(error for error in errors if error["column"] == "Ürün Seviyesi")
        self.assertEqual(issue["code"], "INVALID_ENUM_VALUE")
        self.assertIn("Mamul, Yarı Mamul, Hammadde", issue["message"])

    def test_populated_optional_sheets_are_validated_with_field_locations(self):
        def change(book):
            mapping = book["Malzeme_Tedarikciler"]
            mapping.append(["UNKNOWN", "SUP-1", 0.5, None, None])
            suppliers = book["Tedarikciler"]
            suppliers.append(["SUP-1", "Tedarikçi", 1.2, None, None, None, None, None])
            events = book["Events"]
            events.append([2026, 54, 54, "Örnek Grup", None, "campaign", None, None, None, None])
        _, errors, _, extras = parse_workbook_details(self.workbook(change))
        locations = {(error["sheet"], error["column"]) for error in errors}
        self.assertIn(("Malzeme_Tedarikciler", "Ürün Kodu"), locations)
        self.assertIn(("Tedarikciler", "Sipariş Karşılama Oranı (%)"), locations)
        self.assertIn(("Events", "Başlangıç Hafta"), locations)
        self.assertEqual(extras["v3_supplier_inputs"]["suppliers"][0]["supplier_code"], "SUP-1")

    def test_cross_reference_errors_include_safe_identifier_and_correction_guidance(self):
        def change(book):
            mapping = book["Malzeme_Tedarikciler"]
            mapping.append(["MISSING-SKU", "MISSING-SUPPLIER", 0.5, None, None])
        _, errors, warnings, _ = parse_workbook_details(self.workbook(change))
        product = next(error for error in errors if error["column"] == "Ürün Kodu")
        supplier = next(error for error in errors if error["column"] == "Tedarikçi Kodu")
        self.assertEqual((product["code"], product["severity"], product["sheet"], product["row"]), ("INVALID_REFERENCE", "ERROR", "Malzeme_Tedarikciler", 2))
        self.assertIn("`MISSING-SKU`", product["message"])
        self.assertIn("Temel_Veriler", product["message"])
        self.assertIn("ekleyin veya ilgili satırı", product["message"])
        self.assertIn("`MISSING-SUPPLIER`", supplier["message"])
        self.assertIn("Tedarikciler", supplier["message"])
        self.assertIn("ekleyin veya ilgili satırı", supplier["message"])
        self.assertEqual(warnings, [])

    def test_valid_supplier_references_remain_valid(self):
        def change(book):
            book["Tedarikciler"].append(["SUP-1", "Tedarikçi", 0.9, 0.1, 0.8, 0.1, 7, 1])
            book["Malzeme_Tedarikciler"].append(["SKU-001", "SUP-1", 1, 0, "2026-12-31"])
        _, errors, _, extras = parse_workbook_details(self.workbook(change))
        self.assertEqual(errors, [])
        self.assertEqual(extras["v3_supplier_inputs"]["material_suppliers"][0]["supplier_code"], "SUP-1")

    def test_absent_or_empty_optional_sheets_are_legal(self):
        def change(book):
            book.remove(book["Malzeme_Tedarikciler"])
            book.remove(book["Tedarikciler"])
            book.remove(book["Events"])
        _, errors, _, extras = parse_workbook_details(self.workbook(change))
        self.assertEqual(errors, [])
        self.assertEqual(extras["v3_supplier_inputs"], {"suppliers": [], "material_suppliers": []})
        self.assertEqual(extras["v3_events"], [])
